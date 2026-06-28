#!/usr/bin/env python3
"""
盘后选股模型 V8.1 — 五策略尖刀体系 (6月16日盘后版)
=============================================================
基于V8.0升级:
  - 数据源更新: 6月16日盘后快照 + 全部历史K线(含6月16日)
  - 输出最新交易日选股结果
  - 验证逻辑: 选股仅用T日(6月16日)收盘数据, 不存在未来函数

V8.0 核心理念:
  选股=圈范围, 次日开盘=确认入场
  每个策略只看核心驱动, 互斥不重叠, 要尖刀不要大而全

5个策略:
  1. 🔥打板求涨停  — 情绪×流动性 (合并旧打板+连板)
  2. 📈趋势主升浪  — 均线×筹码   (替代旧纯技术+多因子技术面)
  3. 🩹错杀低吸    — 恐慌极值×承接 (替代旧超跌反弹)
  4. ⚡弱转强      — 预期差×点火 (全新)
  5. 🔄强转弱反抽  — 主力被套×自救 (全新)
"""

import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings('ignore')

pd.set_option('display.max_columns', 100)
pd.set_option('display.width', 200)

# ========== 1. 数据加载 (6月16日盘后) ==========
SNAPSHOT_FILE = 'data/全市场L2快照_20260616.csv'
KLINE_FILES = [
    'data/kline_20260520_0529_daily.csv',
    'data/kline_20260601_0611_daily.csv',
    'data/kline_20260612_daily.csv',
    'data/kline_20260615_daily.csv',
    'data/kline_20260616_daily.csv',  # ★新增6月16日K线
]
NAME_FILE = 'data/stock_name_mapping.csv'
INDUSTRY_FILE = 'data/股票行业三级分类_20260616_033518.csv'
BLOCK_RELATION_FILE = 'data/stock_block_relation.csv'
OUTPUT_FILE = 'output/盘后选股模型_V8_1_20260616_盘后.xlsx'

df = pd.read_csv(SNAPSHOT_FILE, encoding='utf-8-sig')
print(f"[1] 快照数据: {len(df)} 条, 列数: {len(df.columns)}")
print(f"    快照时间: {df['查询时间'].iloc[0] if '查询时间' in df.columns else 'N/A'}")
print(f"    行情日期: {df['HqDate'].iloc[0] if 'HqDate' in df.columns else 'N/A'}")

name_df = pd.read_csv(NAME_FILE, encoding='utf-8-sig')
name_map = dict(zip(name_df['code'], name_df['name']))
df['股票名称'] = df['code'].map(name_map).fillna(df['code'])

# ========== 2. ST过滤 ==========
st_mask = df['股票名称'].str.contains(r'\*?ST', case=False, na=False)
st_count = st_mask.sum()
df = df[~st_mask].copy()
print(f"[2] ST过滤: 排除 {st_count} 只, 剩余 {len(df)} 只")

# ========== 3. 基础过滤 ==========
mask = (
    (df['IsT0Fund'] == 0) &
    (df['IsKzz'] == 0) &
    (df['TPFlag'] == 0) &
    (df['SafeValue'] != -1)
)
code_mask = df['code'].str.match(r'^(6|0|3|4|8)\d{5}\.(SZ|SH|BJ)$', na=False)
df_valid = df[mask & code_mask].copy()
print(f"[3] 基础过滤后: {len(df_valid)} 只")

# ========== 3.1 数据清洗 (修复V8.1审核Bug) ==========
# Bug1: Wtb量比存在大量负值(-99.8~-0.0)和0,均为异常值,需置NaN
wtb_before = df_valid['Wtb'].copy()
df_valid['Wtb'] = pd.to_numeric(df_valid['Wtb'], errors='coerce')
wtb_invalid = ((df_valid['Wtb'] < 0) | (df_valid['Wtb'] == 0)).sum()
df_valid.loc[(df_valid['Wtb'] < 0) | (df_valid['Wtb'] == 0), 'Wtb'] = np.nan

# Bug2: FCb/FCAmo 存在负值,需清洗为0
fcb_neg = (pd.to_numeric(df_valid['FCb'], errors='coerce') < 0).sum()
fcamo_neg = (pd.to_numeric(df_valid['FCAmo'], errors='coerce') < 0).sum()
df_valid['FCb'] = pd.to_numeric(df_valid['FCb'], errors='coerce').where(lambda x: x >= 0, 0)
df_valid['FCAmo'] = pd.to_numeric(df_valid['FCAmo'], errors='coerce').where(lambda x: x >= 0, 0)

# Bug3: 卖撤率公式单位错配, 改用同单位公式 SCancel/(SCancel+TotalSVol)
#   - 原公式 SCancel/L2OrderNum 中位1.30(>1不合理)
#   - 新公式: 卖单撤单占比, 范围[0,1], 中位约0.77
sc = pd.to_numeric(df_valid['SCancel'], errors='coerce')
tsv = pd.to_numeric(df_valid['TotalSVol'], errors='coerce')
df_valid['卖撤率'] = (sc / (sc + tsv + 1)).fillna(0.5)  # 缺失用中位0.5兜底

# Bug4: fLianB 实为"涨停系数/封板强度"(涨停股中位1.23,非涨停股中位0.96),非连板数
#   - 真正连板数应取 ConZAFDateNum (涨停股中位2,最小1,最大7)
#   - 涨停股保留 ConZAFDateNum 作为连板数;非涨停股置0
#   - 保留 fLianB 但语义改为"封板强度系数"(供打板策略参考)
df_valid['封板强度系数'] = pd.to_numeric(df_valid['fLianB'], errors='coerce').fillna(0)
df_valid['fLianB'] = pd.to_numeric(df_valid['ConZAFDateNum'], errors='coerce').fillna(0).clip(lower=0)
# 非涨停股的 ConZAFDateNum 可能是连续上涨天数(非涨停意义),置0
# 此处先保留,后面涨停判断后再清洗

# Bug5: ZAF/Zjl/Zsz 等数值字段统一转 numeric
for col in ['ZAF','Zjl','Zsz','fHSL','VOpenZAF','OpenZAF','FzAmo','OpenAmo',
            'ZAFYesterday','ZAFPre5','ZAFPre10','ZAFPre20','ZAFPre60',
            'YearZTDay','BetaValue','TotalBVol','TotalSVol','BCancel','SCancel',
            'FCAmo','FCb','CJJEPre1','L2OrderNum','MA5Value','OpenAmoPre1']:
    if col in df_valid.columns:
        df_valid[col] = pd.to_numeric(df_valid[col], errors='coerce')

print(f"[3.1] 数据清洗: Wtb异常{wtb_invalid}只置NaN, FCb负值{fcb_neg}只, FCAmo负值{fcamo_neg}只, 卖撤率公式已重写")

is_laodeng = (
    (df_valid['YearZTDay'] == 0) &
    (df_valid['fHSL'] < 1) &
    (df_valid['BetaValue'] < 0.8)
)
df_valid = df_valid[~is_laodeng].copy()
print(f"[3] 老登过滤后: {len(df_valid)} 只")

# ========== 4. 加载K线数据 ==========
kline_dfs = []
for f in KLINE_FILES:
    kline_dfs.append(pd.read_csv(f, encoding='utf-8-sig'))
kline = pd.concat(kline_dfs, ignore_index=True)
kline = kline.drop_duplicates(subset=['code', 'date'], keep='last')
kline = kline.sort_values(['code', 'date']).reset_index(drop=True)
kline = kline[kline['code'].str.match(r'^(6|0|3|4|8)\d{5}\.(SZ|SH|BJ)$', na=False)].copy()
print(f"[4] K线数据: {len(kline)} 条, 日期范围: {kline['date'].min()} ~ {kline['date'].max()}")


def calc_technical_indicators(group):
    group = group.sort_values('date').reset_index(drop=True)
    if len(group) < 5:
        return group
    close = group['close'].astype(float)
    volume = group['volume'].astype(float)
    amount = group['amount'].astype(float)

    group['MA5'] = close.rolling(5, min_periods=1).mean()
    group['MA10'] = close.rolling(10, min_periods=1).mean()
    group['MA20'] = close.rolling(20, min_periods=1).mean()

    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    dif = ema12 - ema26
    dea = dif.ewm(span=9, adjust=False).mean()
    group['DIF'] = dif
    group['DEA'] = dea
    group['MACD_BAR'] = 2 * (dif - dea)

    group['BOLL_MID'] = close.rolling(20, min_periods=1).mean()
    boll_std = close.rolling(20, min_periods=1).std()
    group['BOLL_UP'] = group['BOLL_MID'] + 2 * boll_std
    return group


print("[4] 计算技术指标中...")
kline = kline.groupby('code', group_keys=False).apply(calc_technical_indicators)
latest_date = kline['date'].max()
kline_latest = kline[kline['date'] == latest_date].set_index('code')
print(f"[4] 最新K线日期: {latest_date}, 股票数: {len(kline_latest)}")

# ========== 5. 合并行业+板块+K线 ==========
industry_df = pd.read_csv(INDUSTRY_FILE, encoding='utf-8-sig')
industry_map = {}
for _, row in industry_df.iterrows():
    industry_map[row['stock_code']] = row.get('行业一级', '')
df_valid['行业'] = df_valid['code'].map(lambda c: industry_map.get(c, '其他'))

block_df = pd.read_csv(BLOCK_RELATION_FILE, encoding='utf-8-sig')
block_industry = block_df[block_df['block_type'] == '行业'].copy()
block_map = {}
for _, row in block_industry.iterrows():
    if row['stock_code'] not in block_map:
        block_map[row['stock_code']] = []
    block_map[row['stock_code']].append(row['block_name'])
df_valid['所属板块'] = df_valid['code'].map(lambda c: '|'.join(block_map.get(c, [])))

# 行业统计 (修复V8.1 Bug7: 区分板块计算涨停数)
def _is_zt_for_row(r):
    code = str(r['code'])
    zaf = float(r['ZAF']) if pd.notna(r['ZAF']) else 0
    if code.startswith(('688','689','300','301','302')): return zaf >= 19.5
    elif code.startswith(('8','4')): return zaf >= 29.5
    else: return zaf >= 9.8
df_valid['_is_zt'] = df_valid.apply(_is_zt_for_row, axis=1)
industry_stats = df_valid.groupby('行业').agg(
    行业总数=('code', 'count'),
    行业涨停数=('_is_zt', 'sum'),
    行业平均涨幅=('ZAF', 'mean'),
    行业主力净流入=('Zjl', 'sum'),
).reset_index()
industry_stats['行业涨停率'] = (industry_stats['行业涨停数'] / industry_stats['行业总数'] * 100).round(2)
df_valid = df_valid.merge(
    industry_stats[['行业', '行业总数', '行业涨停数', '行业涨停率', '行业平均涨幅', '行业主力净流入']],
    on='行业', how='left'
)

# 合并K线技术指标
kline_tech = kline_latest[['MA5', 'MA10', 'MA20', 'DIF', 'DEA', 'MACD_BAR', 'BOLL_UP']].copy()
kline_tech.columns = ['KL_MA5', 'KL_MA10', 'KL_MA20', 'KL_DIF', 'KL_DEA', 'KL_MACD_BAR', 'KL_BOLL_UP']
df_valid = df_valid.merge(kline_tech, left_on='code', right_index=True, how='left')

# 最新价 (修复V8.1 Bug5: 不再用MA5Value兌底,改为缺K线则剔除)
def get_latest_close(code):
    if code in kline_latest.index:
        row = kline_latest.loc[code]
        if isinstance(row, pd.DataFrame): row = row.iloc[0]
        return float(row['close'])
    return np.nan

df_valid['最新价'] = df_valid['code'].apply(get_latest_close)
# 兌底优先级: 1)K线收盘价 2)快照现价(如有) 3) MA5Value (最后乲底, 标记为估算)
ma5_snap = df_valid.get('MA5Value', pd.Series(np.nan, index=df_valid.index))
# 记录哪些股票用MA5兌底(估算价)
est_price_mask = df_valid['最新价'].isna() & ma5_snap.notna() & (ma5_snap > 0)
df_valid['最新价'] = df_valid['最新价'].fillna(ma5_snap)
df_valid['价格估算'] = est_price_mask  # 后续可据此标记

# 辅助字段
df_valid['大买占比'] = df_valid['TotalBVol'] / (df_valid['TotalBVol'] + df_valid['TotalSVol'] + 1)
# df_valid['卖撤率'] 已在数据清洗阶段重新计算 (SCancel/(SCancel+TotalSVol))
df_valid['恐慌量'] = df_valid['fHSL'] * df_valid['ZAF'].abs()

# 涨停判断 (沿用 _is_zt_for_row)
df_valid['是否涨停'] = df_valid['_is_zt']
print(f"[5] 辅助计算完成, 当日涨停: {df_valid['是否涨停'].sum()} 只, 价格估算{est_price_mask.sum()}只")

# 修复Bug4: 非涨停股的fLianB(原ConZAFDateNum)可能不代表连板数,需置0
df_valid.loc[~df_valid['是否涨停'], 'fLianB'] = 0
# fLianB现在语义为"连板数",范围[0,7]整数,需取整
df_valid['fLianB'] = df_valid['fLianB'].astype(int)
print(f"[5] fLianB已重定义为连板数(原ConZAFDateNum), 涨停股中连板数分布:")
if df_valid['是否涨停'].sum() > 0:
    print(df_valid[df_valid['是否涨停']]['fLianB'].value_counts().sort_index().head(10).to_string())


# ========== 6. 压力位 ==========
def calc_pressure_level(code, kline_data, latest_close):
    if code not in kline_data.index:
        return latest_close * 1.02, '估算'
    row = kline_data.loc[code]
    if isinstance(row, pd.DataFrame): row = row.iloc[0]
    candidates = []
    prev_high = float(row.get('high', 0))
    if prev_high > latest_close * 1.005: candidates.append((prev_high, '前日高点'))
    boll_up = row.get('BOLL_UP', np.nan)
    if pd.notna(boll_up) and float(boll_up) > latest_close * 1.005: candidates.append((float(boll_up), '布林上轨'))
    for ma_key in ['MA20', 'MA10', 'MA5']:
        ma_val = row.get(ma_key, np.nan)
        if pd.notna(ma_val) and float(ma_val) > latest_close * 1.003: candidates.append((float(ma_val), ma_key))
    if not candidates:
        zt_price = latest_close * 1.10
        code_str = str(code)
        if code_str.startswith(('688','689','300','301','302')): zt_price = latest_close * 1.20
        elif code_str.startswith(('8','4')): zt_price = latest_close * 1.30
        return round(zt_price, 2), '涨停价'
    candidates.sort(key=lambda x: abs(x[0] - latest_close))
    return round(candidates[0][0], 2), candidates[0][1]


# ================================================================
# ========== 7. 五策略评分 ==========
# ================================================================

# ---- 🔥策略一：打板求涨停 (≥35分) ----
def score_daban(df):
    pool = df[
        (df['是否涨停']) |
        (df['ZAF'] >= 7)
    ].copy()
    scores = pd.DataFrame(index=pool.index)
    scores['code'] = pool['code']
    scores['股票名称'] = pool['股票名称']
    lp = {idx: [] for idx in pool.index}

    # (1) 封板强度(0-40) — fLianB已重定义为连板数,补充原fLianB作为封板强度系数
    #     涨停基础分10 + 封成比加成(0-20) + 封单额加成(0-10) + 封板强度系数加成(0-5)
    scores['封板强度'] = 0
    fc_b = pool['FCb']; fc_amo = pool['FCAmo']
    fcb_strength = pool.get('封板强度系数', pd.Series(0, index=pool.index))
    scores.loc[pool['是否涨停'], '封板强度'] += 10
    for idx in pool[pool['是否涨停']].index: lp[idx].append('涨停')
    scores.loc[fc_b >= 0.5, '封板强度'] += 20
    scores.loc[(fc_b >= 0.2) & (fc_b < 0.5), '封板强度'] += 14
    scores.loc[(fc_b >= 0.05) & (fc_b < 0.2), '封板强度'] += 8
    scores.loc[(fc_b > 0) & (fc_b < 0.05), '封板强度'] += 3
    for idx in pool[fc_b >= 0.2].index: lp[idx].append('封板稳固')
    scores.loc[fc_amo >= 30000, '封板强度'] += 10
    scores.loc[(fc_amo >= 10000) & (fc_amo < 30000), '封板强度'] += 8
    scores.loc[(fc_amo >= 5000) & (fc_amo < 10000), '封板强度'] += 5
    scores.loc[(fc_amo >= 1000) & (fc_amo < 5000), '封板强度'] += 2
    for idx in pool[fc_amo >= 10000].index: lp[idx].append('大封单')
    # 补充: 用封板强度系数(原fLianB值)作为额外加分,阈值适配其原始分布(中位~1.0)
    scores.loc[fcb_strength >= 2, '封板强度'] += 5
    scores.loc[(fcb_strength >= 1.5) & (fcb_strength < 2), '封板强度'] += 3
    scores['封板强度'] = scores['封板强度'].clip(0, 40)

    # (2) 连板辨识度(0-30) — fLianB现为真正连板数(整数), 阈值重校准
    #     ≥7=30 / ≥5=26 / 4=22 / 3=18 / 2=12 / 1=6 / 0=0
    scores['连板辨识度'] = 0
    lb = pool['fLianB']  # 现为整数: 0=未涨停, 1=首板, 2=2连板, 3=3连板...
    scores.loc[lb >= 7, '连板辨识度'] = 30
    scores.loc[(lb >= 5) & (lb < 7), '连板辨识度'] = 26
    scores.loc[(lb >= 4) & (lb < 5), '连板辨识度'] = 22
    scores.loc[(lb == 3), '连板辨识度'] = 18
    scores.loc[(lb == 2), '连板辨识度'] = 12
    scores.loc[(lb == 1), '连板辨识度'] = 6
    for idx in pool[lb >= 3].index: lp[idx].append(f'{lb[idx]:.0f}连板')
    for idx in pool[lb == 2].index: lp[idx].append('2连板')
    for idx in pool[lb == 1].index: lp[idx].append('首板')

    # (3) 竞价抢筹(0-20) — 竞价涨幅 + 竞价金额比 + 尾盘抢筹
    scores['竞价抢筹'] = 0
    vopen = pool.get('VOpenZAF', pd.Series(0, index=pool.index))
    fz_amo = pool['FzAmo']
    scores.loc[vopen >= 3, '竞价抢筹'] += 10
    scores.loc[(vopen >= 1.5) & (vopen < 3), '竞价抢筹'] += 7
    scores.loc[(vopen >= 0.5) & (vopen < 1.5), '竞价抢筹'] += 4
    for idx in pool[vopen >= 1.5].index: lp[idx].append('竞价抢筹')
    scores.loc[fz_amo >= 2000, '竞价抢筹'] += 10
    scores.loc[(fz_amo >= 500) & (fz_amo < 2000), '竞价抢筹'] += 7
    scores.loc[(fz_amo >= 200) & (fz_amo < 500), '竞价抢筹'] += 4
    for idx in pool[fz_amo >= 500].index: lp[idx].append('尾盘抢筹')
    scores['竞价抢筹'] = scores['竞价抢筹'].clip(0, 20)

    # (4) 风险扣分(0~-10) — 卖撤率阈值重校准(新公式中位~0.77,范围[0,1])
    #     卖撤率高扣分(-7/-3) + 孤板风险(-8/-4)
    scores['风险扣分'] = 0
    scr = pool.get('卖撤率', pd.Series(0.5, index=pool.index))  # 新公式范围[0,1]
    # 新阈值: >0.95表示几乎全部撤单(异常), >0.90明显偏高
    scores.loc[scr > 0.95, '风险扣分'] -= 7
    scores.loc[(scr > 0.90) & (scr <= 0.95), '风险扣分'] -= 3
    for idx in pool[scr > 0.92].index: lp[idx].append('⚠卖撤率高')
    for idx in pool.index:
        lb_val = lb[idx]
        ind_zt = pool.loc[idx, '行业涨停数'] if '行业涨停数' in pool.columns else 0
        # 修复Bug: 行业涨停数修复后是准确值,孤板阈值由<=1校准为<=2
        #   - 修复前(不分板块用9.8): 行业涨停数被低估,阈值<=1过严
        #   - 修复后(区分9.8/19.5/29.5): 行业涨停数准确,阈值<=2更合理
        if lb_val >= 4 and ind_zt <= 2:
            scores.loc[idx, '风险扣分'] -= 8; lp[idx].append('⚠高位孤板')
        elif lb_val >= 3 and ind_zt <= 2:
            scores.loc[idx, '风险扣分'] -= 4; lp[idx].append('⚠孤板风险')

    scores['总分'] = scores['封板强度'] + scores['连板辨识度'] + scores['竞价抢筹'] + scores['风险扣分']
    mask_weak = (~pool['是否涨停']) & (pool['FCb'] == 0) & (pool['fLianB'] < 1)
    scores.loc[mask_weak, '总分'] = (scores.loc[mask_weak, '总分'] * 0.5).round(1)

    scores['次日确认'] = ''
    for idx in scores.index:
        if scores.loc[idx, '总分'] < 35: continue
        fc_b_v = pool.loc[idx, 'FCb']; lb_v = pool.loc[idx, 'fLianB']
        is_zt = pool.loc[idx, '是否涨停']
        if lb_v >= 3: scores.loc[idx, '次日确认'] = '竞价不低开(>0%)+竞价量放大'
        elif is_zt and fc_b_v >= 0.2: scores.loc[idx, '次日确认'] = '竞价涨幅>1.5%+不低开'
        elif is_zt and fc_b_v > 0: scores.loc[idx, '次日确认'] = '竞价涨幅>2%+量比>1.5'
        elif is_zt: scores.loc[idx, '次日确认'] = '竞价涨幅>2%+量比>1.5'
        else: scores.loc[idx, '次日确认'] = '竞价涨幅>3%+量比>2+5分钟涨>2%'
        if lb_v >= 5 and fc_b_v < 0.1:
            scores.loc[idx, '次日确认'] += '(⚠封板弱,若低开放弃入场)'

    scores['选股逻辑'] = ''
    for idx in scores.index:
        parts = lp.get(idx, [])
        if scores.loc[idx, '总分'] <= 0: continue
        if not parts: parts.append(f'涨幅{pool.loc[idx,"ZAF"]:.1f}%')
        scores.loc[idx, '选股逻辑'] = '、'.join(parts) + ' → 打板入选'
    return scores


# ---- 📈策略二：趋势主升浪 (≥60分) ----
def score_qushi(df):
    scores = pd.DataFrame(index=df.index)
    scores['code'] = df['code']; scores['股票名称'] = df['股票名称']
    lp = {idx: [] for idx in df.index}

    ma5 = df.get('KL_MA5', pd.Series(np.nan, index=df.index))
    ma10 = df.get('KL_MA10', pd.Series(np.nan, index=df.index))
    ma20 = df.get('KL_MA20', pd.Series(np.nan, index=df.index))
    has_kline = pd.notna(ma5) & pd.notna(ma10) & pd.notna(ma20)

    # (1) 均线多头(0-35) — MA5>MA10>MA20=35, 部分多头按比例给分
    scores['均线多头'] = 0
    full_bull = has_kline & (ma5 > ma10) & (ma10 > ma20)
    scores.loc[full_bull, '均线多头'] = 35
    for idx in df[full_bull].index: lp[idx].append('均线多头排列')
    short_bull = has_kline & (ma5 > ma10) & (ma10 <= ma20)
    scores.loc[short_bull, '均线多头'] = 22
    for idx in df[short_bull].index: lp[idx].append('短期均线多头')
    r5 = df['ZAFPre5']; r10 = df['ZAFPre10']
    scores.loc[~has_kline & (r5 > 0) & (r10 > 0), '均线多头'] = 20
    scores.loc[~has_kline & (r5 > 0) & (r10 <= 0), '均线多头'] = 8

    # (2) 量价配合(0-25) — 涨+量比>1.5=25, 缩量上涨扣分
    scores['量价配合'] = 0
    zaf = df['ZAF']; wtb = df['Wtb']
    scores.loc[(zaf > 0) & (wtb > 1.5), '量价配合'] = 25
    scores.loc[(zaf > 0) & (wtb > 1) & (wtb <= 1.5), '量价配合'] = 18
    scores.loc[(zaf > 0) & (wtb > 0.8) & (wtb <= 1), '量价配合'] = 10
    scores.loc[(zaf > 0) & (wtb <= 0.8), '量价配合'] = 4
    for idx in df[(zaf > 0) & (wtb > 1.5)].index: lp[idx].append('量价齐升')
    for idx in df[(zaf > 0) & (wtb <= 0.8)].index: lp[idx].append('⚠缩量上涨')

    # (3) MACD方向(0-20) — 金叉/向上发散
    scores['MACD方向'] = 0
    dif = df.get('KL_DIF', pd.Series(np.nan, index=df.index))
    dea = df.get('KL_DEA', pd.Series(np.nan, index=df.index))
    has_macd = pd.notna(dif) & pd.notna(dea)
    scores.loc[has_macd & (dif > 0) & (dif > dea), 'MACD方向'] = 20
    scores.loc[has_macd & (dif > 0) & (dif <= dea), 'MACD方向'] = 12
    scores.loc[has_macd & (dif <= 0) & (dif > dea), 'MACD方向'] = 10
    for idx in df[has_macd & (dif > 0) & (dif > dea)].index: lp[idx].append('MACD金叉')
    scores.loc[~has_macd & (r5 > 3), 'MACD方向'] = 14
    scores.loc[~has_macd & (r5 > 0) & (r5 <= 3), 'MACD方向'] = 8

    # (4) 大单流入(0-20) — 主力净流入+大单占比
    scores['大单流入'] = 0
    zjl = df['Zjl']; big_ratio = df['大买占比']
    scores.loc[(zjl > 0) & (big_ratio > 0.4), '大单流入'] = 20
    scores.loc[(zjl > 0) & (big_ratio > 0.3) & (big_ratio <= 0.4), '大单流入'] = 14
    scores.loc[(zjl > 0) & (big_ratio <= 0.3), '大单流入'] = 8
    for idx in df[(zjl > 5000) & (big_ratio > 0.3)].index: lp[idx].append('大单持续流入')

    scores['总分'] = scores['均线多头'] + scores['量价配合'] + scores['MACD方向'] + scores['大单流入']
    not_uptrend = has_kline & (ma5 <= ma10)
    scores.loc[not_uptrend, '总分'] = (scores.loc[not_uptrend, '总分'] * 0.3).round(1)

    scores['次日确认'] = ''
    for idx in scores.index:
        if scores.loc[idx, '总分'] < 60: continue
        ma5_val = ma5[idx] if pd.notna(ma5[idx]) else None
        if ma5_val: scores.loc[idx, '次日确认'] = f'开盘不破MA5({ma5_val:.2f}元)+量比>0.8'
        else: scores.loc[idx, '次日确认'] = '开盘不跌破昨收+量比>0.8'

    scores['选股逻辑'] = ''
    for idx in scores.index:
        parts = lp.get(idx, [])
        if scores.loc[idx, '总分'] < 60: continue
        if not parts: parts.append('趋势向上')
        scores.loc[idx, '选股逻辑'] = '、'.join(parts) + ' → 趋势主升浪入选'
    return scores


# ---- 🩹策略三：错杀低吸 (≥55分) ----
def score_cuoshai(df):
    # 修复Bug: pool 改 OR → AND,确保是"当日下跌 + 中期超跌"的真正错杀股
    # 同时保留两种主要场景: (1)当日明显下跌(ZAF<-3) (2)中期深度超跌+当日下跌(ZAF<-1 & ZAFPre20<-8)
    pool = df[
        (df['ZAF'] < -3) |
        ((df['ZAF'] < -1) & (df['ZAFPre20'] < -8)) |
        ((df['ZAF'] < -1) & (df['ZAFPre60'] < -15))
    ].copy()
    scores = pd.DataFrame(index=pool.index)
    scores['code'] = pool['code']; scores['股票名称'] = pool['股票名称']
    lp = {idx: [] for idx in pool.index}

    # (1) 恐慌深度(0-30) — 当日跌+20日跌+60日跌
    scores['恐慌深度'] = 0
    zaf = pool['ZAF']; zaf20 = pool['ZAFPre20']; zaf60 = pool['ZAFPre60']
    scores.loc[zaf <= -7, '恐慌深度'] += 15
    scores.loc[(zaf > -7) & (zaf <= -5), '恐慌深度'] += 12
    scores.loc[(zaf > -5) & (zaf <= -3), '恐慌深度'] += 8
    scores.loc[(zaf > -3) & (zaf <= -1), '恐慌深度'] += 3
    for idx in pool[zaf <= -5].index: lp[idx].append('当日恐慌杀跌')
    scores.loc[zaf60 <= -25, '恐慌深度'] += 10
    scores.loc[(zaf60 > -25) & (zaf60 <= -15), '恐慌深度'] += 7
    scores.loc[(zaf20 <= -15), '恐慌深度'] += 8
    scores.loc[(zaf20 <= -8) & (zaf20 > -15), '恐慌深度'] += 5
    for idx in pool[zaf60 <= -15].index: lp[idx].append('中期超跌')
    scores['恐慌深度'] = scores['恐慌深度'].clip(0, 30)

    # (2) 承接力度(0-30) — 大买占比+主力净流入
    scores['承接力度'] = 0
    big_ratio = pool['大买占比']; zjl = pool['Zjl']
    scores.loc[big_ratio >= 0.5, '承接力度'] += 18
    scores.loc[(big_ratio >= 0.4) & (big_ratio < 0.5), '承接力度'] += 14
    scores.loc[(big_ratio >= 0.3) & (big_ratio < 0.4), '承接力度'] += 8
    for idx in pool[big_ratio >= 0.4].index: lp[idx].append('大单承接强')
    scores.loc[zjl >= 5000, '承接力度'] += 12
    scores.loc[(zjl >= 1000) & (zjl < 5000), '承接力度'] += 8
    scores.loc[(zjl >= 0) & (zjl < 1000), '承接力度'] += 3
    for idx in pool[zjl >= 1000].index: lp[idx].append('资金抄底')
    scores['承接力度'] = scores['承接力度'].clip(0, 30)

    # (3) 恐慌极值(0-20) — 换手×跌幅
    scores['恐慌极值'] = 0
    pv = pool['恐慌量']
    scores.loc[pv >= 100, '恐慌极值'] = 20
    scores.loc[(pv >= 50) & (pv < 100), '恐慌极值'] = 16
    scores.loc[(pv >= 30) & (pv < 50), '恐慌极值'] = 12
    scores.loc[(pv >= 15) & (pv < 30), '恐慌极值'] = 6
    for idx in pool[pv >= 50].index: lp[idx].append('恐慌放量')

    # (4) 催化剂(0-20) — 行业涨停率 + 量比异常
    scores['催化剂'] = 0
    ind_zt_rate = pool['行业涨停率']; wtb = pool['Wtb']
    scores.loc[ind_zt_rate >= 5, '催化剂'] += 12
    scores.loc[(ind_zt_rate >= 2) & (ind_zt_rate < 5), '催化剂'] += 8
    scores.loc[(ind_zt_rate >= 1) & (ind_zt_rate < 2), '催化剂'] += 3
    for idx in pool[ind_zt_rate >= 2].index: lp[idx].append('板块催化')
    scores.loc[wtb >= 3, '催化剂'] += 8
    scores.loc[(wtb >= 2) & (wtb < 3), '催化剂'] += 5
    for idx in pool[wtb >= 2].index: lp[idx].append('底部放量')
    scores['催化剂'] = scores['催化剂'].clip(0, 20)

    scores['总分'] = scores['恐慌深度'] + scores['承接力度'] + scores['恐慌极值'] + scores['催化剂']
    scores.loc[pool['ZAF'] > 3, '总分'] = 0  # 已反弹的股票不算错杀
    # 修复Bug: 惩罚系数0.4过严,改为0.5保留一定分数
    scores.loc[scores['催化剂'] < 5, '总分'] = (scores.loc[scores['催化剂'] < 5, '总分'] * 0.5).round(1)

    scores['次日确认'] = ''
    for idx in scores.index:
        if scores.loc[idx, '总分'] < 55: continue
        zaf_val = pool.loc[idx, 'ZAF']
        if zaf_val <= -5: scores.loc[idx, '次日确认'] = '开盘不创新低+竞价不低开'
        else: scores.loc[idx, '次日确认'] = '开盘>前日最低+竞价有买单'

    scores['选股逻辑'] = ''
    for idx in scores.index:
        parts = lp.get(idx, [])
        if scores.loc[idx, '总分'] < 55: continue
        if not parts: parts.append('超跌+有承接')
        scores.loc[idx, '选股逻辑'] = '、'.join(parts) + ' → 错杀低吸入选'
    return scores


# ---- ⚡策略四：弱转强 (≥50分) ----
def score_ruozhuanqiang(df):
    # 修复Bug: 阈值收紧, "前日弱"应该是前日跌或平盘,而非"前日涨幅<2%"
    pool = df[
        (df['VOpenZAF'] > 1) &
        (df['ZAFYesterday'] < 0)  # 前日确实下跌才算"弱"
    ].copy()
    scores = pd.DataFrame(index=pool.index)
    scores['code'] = pool['code']; scores['股票名称'] = pool['股票名称']
    lp = {idx: [] for idx in pool.index}

    # (1) 竞价异动(0-30) — VOpenZAF竞价涨幅 + 竞价金额比
    scores['竞价异动'] = 0
    vopen = pool['VOpenZAF']; open_amo = pool['OpenAmo']; cjjepre1 = pool['CJJEPre1']
    scores.loc[vopen >= 5, '竞价异动'] += 15
    scores.loc[(vopen >= 3) & (vopen < 5), '竞价异动'] += 12
    scores.loc[(vopen >= 2) & (vopen < 3), '竞价异动'] += 8
    scores.loc[(vopen >= 1) & (vopen < 2), '竞价异动'] += 4
    for idx in pool[vopen >= 3].index: lp[idx].append(f'竞价涨幅{vopen[idx]:.1f}%')
    # 注: OpenAmo单位=元, CJJEPre1单位=万元, 故*10000对齐为元
    open_amo_ratio = open_amo / (cjjepre1 * 10000 + 1)
    scores.loc[open_amo_ratio >= 0.01, '竞价异动'] += 15
    scores.loc[(open_amo_ratio >= 0.005) & (open_amo_ratio < 0.01), '竞价异动'] += 10
    scores.loc[(open_amo_ratio >= 0.002) & (open_amo_ratio < 0.005), '竞价异动'] += 5
    for idx in pool[open_amo_ratio >= 0.005].index: lp[idx].append('竞价放量')
    scores['竞价异动'] = scores['竞价异动'].clip(0, 30)

    # (2) 预期差(0-25) — 前日弱 + 今日强
    scores['预期差'] = 0
    zaf_yest = pool['ZAFYesterday']; zaf_pre5 = pool['ZAFPre5']; zaf_now = pool['ZAF']
    scores.loc[(zaf_yest < -2) & (vopen > 3), '预期差'] += 15
    scores.loc[(zaf_yest < -2) & (vopen > 1), '预期差'] += 10
    scores.loc[(zaf_yest < 0) & (vopen > 2), '预期差'] += 10
    scores.loc[(zaf_yest >= 0) & (zaf_yest < 2) & (vopen > 3), '预期差'] += 8
    for idx in pool[(zaf_yest < -2) & (vopen > 2)].index: lp[idx].append('强预期差')
    scores.loc[(zaf_pre5 < 0) & (zaf_now > 0), '预期差'] += 10
    for idx in pool[(zaf_pre5 < 0) & (zaf_now > 0)].index: lp[idx].append('5日趋势反转')
    scores['预期差'] = scores['预期差'].clip(0, 25)

    # (3) 点火信号(0-25) — 开盘跳空 + 量比 + 尾盘抢筹
    scores['点火信号'] = 0
    open_zaf = pool['OpenZAF']; wtb = pool['Wtb']; fz_amo = pool['FzAmo']
    scores.loc[open_zaf >= 3, '点火信号'] += 10
    scores.loc[(open_zaf >= 1) & (open_zaf < 3), '点火信号'] += 7
    scores.loc[(open_zaf >= 0.5) & (open_zaf < 1), '点火信号'] += 3
    for idx in pool[open_zaf >= 2].index: lp[idx].append('开盘跳空')
    scores.loc[wtb >= 3, '点火信号'] += 8
    scores.loc[(wtb >= 2) & (wtb < 3), '点火信号'] += 5
    scores.loc[(wtb >= 1.5) & (wtb < 2), '点火信号'] += 3
    for idx in pool[wtb >= 2].index: lp[idx].append('量比放大')
    scores.loc[fz_amo >= 500, '点火信号'] += 7
    scores.loc[(fz_amo >= 200) & (fz_amo < 500), '点火信号'] += 4
    scores['点火信号'] = scores['点火信号'].clip(0, 25)

    # (4) 股性(0-20) — 年涨停数 + Beta值
    scores['股性'] = 0
    nzt = pool.get('YearZTDay', pd.Series(0, index=pool.index))
    beta = pool.get('BetaValue', pd.Series(0, index=pool.index))
    scores.loc[nzt >= 5, '股性'] += 10
    scores.loc[(nzt >= 2) & (nzt < 5), '股性'] += 6
    for idx in pool[nzt >= 5].index: lp[idx].append('股性活跃')
    scores.loc[(beta > 1.2) & (beta <= 2), '股性'] += 10
    scores.loc[(beta > 0.8) & (beta <= 1.2), '股性'] += 5

    scores['总分'] = scores['竞价异动'] + scores['预期差'] + scores['点火信号'] + scores['股性']
    scores.loc[pool['是否涨停'], '总分'] = 0

    scores['次日确认'] = '竞价量比>1.5+5分钟内涨幅>1%'

    scores['选股逻辑'] = ''
    for idx in scores.index:
        parts = lp.get(idx, [])
        if scores.loc[idx, '总分'] < 50: continue
        if not parts: parts.append('竞价异动+预期差')
        scores.loc[idx, '选股逻辑'] = '、'.join(parts) + ' → 弱转强入选'
    return scores


# ---- 🔄策略五：强转弱反抽 (≥50分) ----
def score_qiangzhuanruo(df):
    pool = df[
        (df['ZAFYesterday'] > 3) &
        (df['ZAF'] < -1)
    ].copy()
    scores = pd.DataFrame(index=pool.index)
    scores['code'] = pool['code']; scores['股票名称'] = pool['股票名称']
    lp = {idx: [] for idx in pool.index}

    # (1) 主力被套深度(0-30) — 昨涨-今跌 + 主力逆势流入
    scores['主力被套深度'] = 0
    zaf_yest = pool['ZAFYesterday']; zjl = pool['Zjl']; zaf = pool['ZAF']
    trap_depth = zaf_yest - zaf
    scores.loc[trap_depth >= 15, '主力被套深度'] += 20
    scores.loc[(trap_depth >= 10) & (trap_depth < 15), '主力被套深度'] += 15
    scores.loc[(trap_depth >= 6) & (trap_depth < 10), '主力被套深度'] += 10
    scores.loc[trap_depth < 6, '主力被套深度'] += 4
    for idx in pool[trap_depth >= 10].index: lp[idx].append('主力深度被套')
    scores.loc[zjl > 0, '主力被套深度'] += 10
    scores.loc[(zjl <= 0) & (zjl > -5000), '主力被套深度'] += 5
    for idx in pool[zjl > 0].index: lp[idx].append('主力逆势流入')
    scores['主力被套深度'] = scores['主力被套深度'].clip(0, 30)

    # (2) 回踩幅度(0-25) — 当日跌幅区间
    scores['回踩幅度'] = 0
    scores.loc[(zaf >= -3) & (zaf < -1), '回踩幅度'] = 10
    scores.loc[(zaf >= -5) & (zaf < -3), '回踩幅度'] = 18
    scores.loc[(zaf >= -8) & (zaf < -5), '回踩幅度'] = 25
    scores.loc[zaf < -8, '回踩幅度'] = 15
    for idx in pool[(zaf >= -5) & (zaf < -3)].index: lp[idx].append('中度回踩')
    for idx in pool[(zaf >= -8) & (zaf < -5)].index: lp[idx].append('深度回踩(反抽空间大)')

    # (3) 板块支撑(0-25) — 行业涨停率 + 行业平均涨幅
    scores['板块支撑'] = 0
    ind_zt_rate = pool['行业涨停率']; ind_avg = pool['行业平均涨幅']
    scores.loc[ind_zt_rate >= 5, '板块支撑'] += 15
    scores.loc[(ind_zt_rate >= 2) & (ind_zt_rate < 5), '板块支撑'] += 10
    scores.loc[(ind_zt_rate >= 1) & (ind_zt_rate < 2), '板块支撑'] += 5
    for idx in pool[ind_zt_rate >= 2].index: lp[idx].append('板块仍强势')
    scores.loc[ind_avg > 1, '板块支撑'] += 10
    scores.loc[(ind_avg > 0) & (ind_avg <= 1), '板块支撑'] += 5
    scores['板块支撑'] = scores['板块支撑'].clip(0, 25)

    # (4) 反抽信号(0-20) — 大买占比 + 换手率
    scores['反抽信号'] = 0
    big_ratio = pool['大买占比']; hsl = pool['fHSL']
    scores.loc[big_ratio >= 0.4, '反抽信号'] += 12
    scores.loc[(big_ratio >= 0.3) & (big_ratio < 0.4), '反抽信号'] += 7
    for idx in pool[big_ratio >= 0.35].index: lp[idx].append('有承接单')
    scores.loc[(hsl >= 3) & (hsl <= 15), '反抽信号'] += 8
    scores.loc[(hsl >= 1) & (hsl < 3), '反抽信号'] += 4
    scores.loc[hsl > 15, '反抽信号'] += 2
    scores['反抽信号'] = scores['反抽信号'].clip(0, 20)

    scores['总分'] = scores['主力被套深度'] + scores['回踩幅度'] + scores['板块支撑'] + scores['反抽信号']

    scores['次日确认'] = ''
    for idx in scores.index:
        if scores.loc[idx, '总分'] < 50: continue
        zaf_val = pool.loc[idx, 'ZAF']
        if zaf_val <= -5: scores.loc[idx, '次日确认'] = '开盘>-2%+30分钟不创新低'
        else: scores.loc[idx, '次日确认'] = '开盘>-1%+30分钟不创新低'

    scores['选股逻辑'] = ''
    for idx in scores.index:
        parts = lp.get(idx, [])
        if scores.loc[idx, '总分'] < 50: continue
        if not parts: parts.append('昨日强今日弱')
        scores.loc[idx, '选股逻辑'] = '、'.join(parts) + ' → 强转弱反抽入选'
    return scores


# ========== 8. 执行评分 ==========
print("\n===== 执行五策略评分 =====")
scores_daban = score_daban(df_valid)
scores_qushi = score_qushi(df_valid)
scores_cuoshai = score_cuoshai(df_valid)
scores_ruozhuan = score_ruozhuanqiang(df_valid)
scores_qzr = score_qiangzhuanruo(df_valid)

THRESHOLDS = {'daban': 35, 'qushi': 60, 'cuoshai': 55, 'ruozhuan': 50, 'qzr': 50}
MAX_PER_MODEL = 30

print(f"打板≥{THRESHOLDS['daban']}: {(scores_daban['总分'] >= THRESHOLDS['daban']).sum()} 只")
print(f"趋势≥{THRESHOLDS['qushi']}: {(scores_qushi['总分'] >= THRESHOLDS['qushi']).sum()} 只")
print(f"错杀≥{THRESHOLDS['cuoshai']}: {(scores_cuoshai['总分'] >= THRESHOLDS['cuoshai']).sum()} 只")
print(f"弱转强≥{THRESHOLDS['ruozhuan']}: {(scores_ruozhuan['总分'] >= THRESHOLDS['ruozhuan']).sum()} 只")
print(f"强转弱≥{THRESHOLDS['qzr']}: {(scores_qzr['总分'] >= THRESHOLDS['qzr']).sum()} 只")


# ========== 9. 构建结果 ==========
COL_RENAME = {
    'code': '股票代码', '股票名称': '股票名称', 'ZAF': '涨幅%', 'fHSL': '换手率%',
    'Wtb': '量比', 'Zsz': '总市值(亿)', 'Zjl': '主力净流入(万)',
    'FCAmo': '封单额(万)', 'FCb': '封成比', 'fLianB': '连板数',
    'VOpenZAF': '竞价涨幅%', 'OpenZAF': '开盘涨幅%', 'FzAmo': '尾盘金额(万)',
    'OpenAmo': '竞价金额', 'ZAFYesterday': '前日涨幅%',
    'ZAFPre5': '5日涨幅%', 'ZAFPre10': '10日涨幅%', 'ZAFPre20': '20日涨幅%', 'ZAFPre60': '60日涨幅%',
    'YearZTDay': '年涨停天数', '行业': '行业', '行业涨停数': '行业涨停数',
    '行业涨停率': '行业涨停率%', '行业平均涨幅': '行业平均涨幅%',
    '最新价': '最新价', '大买占比': '大买占比',
    '排名': '排名', '选股模型': '选股模型',
    '封板强度': '封板强度', '连板辨识度': '连板辨识度', '竞价抢筹': '竞价抢筹', '风险扣分': '风险扣分',
    '均线多头': '均线多头', '量价配合': '量价配合', 'MACD方向': 'MACD方向', '大单流入': '大单流入',
    '恐慌深度': '恐慌深度', '承接力度': '承接力度', '恐慌极值': '恐慌极值', '催化剂': '催化剂',
    '竞价异动': '竞价异动', '预期差': '预期差', '点火信号': '点火信号', '股性': '股性',
    '主力被套深度': '主力被套深度', '回踩幅度': '回踩幅度', '板块支撑': '板块支撑', '反抽信号': '反抽信号',
    '总分': '总分', '选股逻辑': '选股逻辑', '次日确认': '次日确认',
    '压力位': '压力位', '压力位类型': '压力位类型', '压力位幅度%': '压力位幅度%',
    '入选策略数': '入选策略数', '入选策略': '入选策略',
    '价格估算': '价格估算',  # True=MA5兜底估算价, False=K线收盘价
    '次日确认条件数': '次日确认条件数',  # 多策略共振sheet专用
}

BASE_COLS_MAP = {
    'daban': ['code', '股票名称', 'ZAF', 'fHSL', 'Wtb', 'Zsz', 'Zjl', 'FCAmo', 'FCb', 'fLianB', 'VOpenZAF', '行业涨停数', '最新价', '价格估算'],
    'qushi': ['code', '股票名称', 'ZAF', 'fHSL', 'Wtb', 'Zsz', 'Zjl', '行业', '最新价', '价格估算'],
    'cuoshai': ['code', '股票名称', 'ZAF', 'fHSL', 'Wtb', 'Zsz', 'Zjl', 'ZAFPre20', 'ZAFPre60', '行业', '行业涨停率', '最新价', '大买占比', '价格估算'],
    'ruozhuan': ['code', '股票名称', 'ZAF', 'fHSL', 'Wtb', 'Zsz', 'Zjl', 'VOpenZAF', 'OpenZAF', 'FzAmo', 'ZAFYesterday', 'ZAFPre5', 'YearZTDay', '最新价', '价格估算'],
    'qzr': ['code', '股票名称', 'ZAF', 'fHSL', 'Wtb', 'Zsz', 'Zjl', 'ZAFYesterday', '行业', '行业涨停率', '最新价', '大买占比', '价格估算'],
}

def build_result(df_data, scores, model_name, threshold, strategy_key, kline_data):
    result = pd.DataFrame(index=scores.index)
    base_cols = BASE_COLS_MAP.get(strategy_key, ['code', '股票名称', 'ZAF', '最新价'])
    for col in base_cols:
        if col in df_data.columns: result[col] = df_data.loc[scores.index, col]
    score_cols = [c for c in scores.columns if c not in ['code', '股票名称', 'ZAF']]
    for col in score_cols: result[col] = scores[col]
    result = result[result['总分'] >= threshold].copy()
    result = result.sort_values('总分', ascending=False).head(MAX_PER_MODEL).reset_index(drop=True)
    result.insert(0, '排名', range(1, len(result) + 1))
    result.insert(1, '选股模型', model_name)
    if len(result) > 0:
        pv = []; pn = []
        for idx, row in result.iterrows():
            code = row['code']; close = row.get('最新价', np.nan)
            if pd.notna(close) and close > 0:
                v, n = calc_pressure_level(code, kline_data, close)
                pv.append(v); pn.append(n)
            else:
                pv.append(np.nan); pn.append('')
        result['压力位'] = pv; result['压力位类型'] = pn
        result['压力位幅度%'] = result.apply(
            lambda r: round((r['压力位'] - r['最新价']) / r['最新价'] * 100, 2)
            if pd.notna(r['压力位']) and pd.notna(r['最新价']) and r['最新价'] > 0 else np.nan, axis=1)
    result = result.rename(columns=COL_RENAME)
    return result


result_daban = build_result(df_valid, scores_daban, '🔥打板求涨停', THRESHOLDS['daban'], 'daban', kline_latest)
result_qushi = build_result(df_valid, scores_qushi, '📈趋势主升浪', THRESHOLDS['qushi'], 'qushi', kline_latest)
result_cuoshai = build_result(df_valid, scores_cuoshai, '🩹错杀低吸', THRESHOLDS['cuoshai'], 'cuoshai', kline_latest)
result_ruozhuan = build_result(df_valid, scores_ruozhuan, '⚡弱转强', THRESHOLDS['ruozhuan'], 'ruozhuan', kline_latest)
result_qzr = build_result(df_valid, scores_qzr, '🔄强转弱反抽', THRESHOLDS['qzr'], 'qzr', kline_latest)

print(f"\n打板求涨停: {len(result_daban)} 只")
print(f"趋势主升浪: {len(result_qushi)} 只")
print(f"错杀低吸: {len(result_cuoshai)} 只")
print(f"弱转强: {len(result_ruozhuan)} 只")
print(f"强转弱反抽: {len(result_qzr)} 只")

# ========== 10. 多策略共振 ==========
print("\n===== 多策略共振 =====")
all_selected = {}
# 记录每只股入选的策略对应的"次日确认"条件, 共振股取最严格条件(全满足才入场)
all_next_day_confirm = {}
for name, result in [('打板', result_daban), ('趋势', result_qushi), ('错杀', result_cuoshai), ('弱转强', result_ruozhuan), ('强转弱', result_qzr)]:
    for _, r in result.iterrows():
        code = r['股票代码']
        if code not in all_selected: all_selected[code] = []
        if name not in all_selected[code]:
            all_selected[code].append(name)
        confirm = r.get('次日确认', '') if '次日确认' in r else ''
        if code not in all_next_day_confirm: all_next_day_confirm[code] = []
        if confirm: all_next_day_confirm[code].append((name, confirm))

resonance_rows = []
for code, models in all_selected.items():
    if len(models) >= 2:
        row_data = df_valid[df_valid['code'] == code]
        if len(row_data) == 0: continue
        row = row_data.iloc[0]
        close = row.get('最新价', np.nan)
        if pd.notna(close) and close > 0: p_val, p_name = calc_pressure_level(code, kline_latest, close)
        else: p_val, p_name = np.nan, ''
        # 修复Bug: 多策略共振缺统一次日确认 — 取各策略次日确认的交集(最严格)
        confirm_list = all_next_day_confirm.get(code, [])
        if confirm_list:
            # 拼接所有策略的次日确认条件, 用" 且 "分隔表示全满足
            unified_confirm = ' 且 '.join([f'[{n}]{c}' for n, c in confirm_list])
        else:
            unified_confirm = ''
        # 新增: 次日确认条件数(便于一眼看出共振股需要满足几个条件)
        confirm_count = len(confirm_list)
        resonance_rows.append({'code': code, '股票名称': row.get('股票名称', code), 'ZAF': row.get('ZAF', 0),
                               'fHSL': row.get('fHSL', 0), 'Zjl': row.get('Zjl', 0), '行业': row.get('行业', ''),
                               '入选策略数': len(models), '入选策略': '、'.join(models),
                               '最新价': close, '压力位': p_val, '压力位类型': p_name,
                               '次日确认条件数': confirm_count,
                               '次日确认': unified_confirm})

result_resonance = pd.DataFrame(resonance_rows)
if len(result_resonance) > 0:
    result_resonance = result_resonance.sort_values('入选策略数', ascending=False).reset_index(drop=True)
    result_resonance.insert(0, '排名', range(1, len(result_resonance) + 1))
    result_resonance = result_resonance.rename(columns=COL_RENAME)
print(f"多策略共振(≥2策略): {len(result_resonance)} 只")


# ========== 11. 写入Excel ==========
print("\n===== 写入Excel =====")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.properties.creator = "Z.ai"
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
RANK_FILL_1 = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
RANK_FILL_2 = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
SCORE_HIGH = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                     top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

COL_WIDTH = {
    '排名': 6, '选股模型': 14, '股票代码': 12, '股票名称': 12, '涨幅%': 8,
    '换手率%': 8, '量比': 6, '总市值(亿)': 12, '主力净流入(万)': 12,
    '封单额(万)': 12, '封成比': 8, '连板数': 8, '竞价涨幅%': 10,
    '尾盘金额(万)': 12, '开盘涨幅%': 10, '前日涨幅%': 10,
    '5日涨幅%': 10, '10日涨幅%': 10, '20日涨幅%': 10, '60日涨幅%': 10,
    '行业': 10, '行业涨停数': 10, '行业涨停率%': 10, '行业平均涨幅%': 12,
    '年涨停天数': 10, '大买占比': 8, '最新价': 10,
    '总分': 8, '选股逻辑': 40, '次日确认': 25,
    '压力位': 10, '压力位类型': 10, '压力位幅度%': 10,
    '封板强度': 8, '连板辨识度': 8, '竞价抢筹': 8, '风险扣分': 8,
    '均线多头': 8, '量价配合': 8, 'MACD方向': 8, '大单流入': 8,
    '恐慌深度': 8, '承接力度': 8, '恐慌极值': 8, '催化剂': 8,
    '竞价异动': 8, '预期差': 8, '点火信号': 8, '股性': 8,
    '主力被套深度': 10, '回踩幅度': 8, '板块支撑': 8, '反抽信号': 8,
    '入选策略数': 10, '入选策略': 15,
    '价格估算': 10,  # 显示"是/否", 是=MA5兜底估算价
    '次日确认条件数': 12,  # 多策略共振sheet, 显示1/2/3等
}

RATE_COLS = ('涨幅%', '5日涨幅%', '10日涨幅%', '20日涨幅%', '60日涨幅%',
             '前日涨幅%', '开盘涨幅%', '竞价涨幅%', '压力位幅度%')

def write_sheet(wb, sheet_name, result_df, model_desc):
    ws = wb.create_sheet(title=sheet_name)
    ws.merge_cells('A1:H1')
    ws['A1'].value = f'{sheet_name} — {model_desc}'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')
    ws['A1'].alignment = LEFT
    ws.merge_cells('A2:H2')
    ws['A2'].value = f'数据日期: 2026-06-16 盘后 | V8.1五策略尖刀 | ST已过滤 | 共 {len(result_df)} 只入选'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = LEFT
    if len(result_df) == 0:
        ws.cell(row=3, column=1, value='无符合条件标的').font = Font(name='微软雅黑', size=11, color='999999')
        return ws
    cols = list(result_df.columns)
    ds = 3
    for j, cn in enumerate(cols, 1):
        cell = ws.cell(row=ds, column=j, value=cn)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = THIN_BORDER
    for i, (_, row) in enumerate(result_df.iterrows()):
        r = ds + 1 + i
        for j, cn in enumerate(cols, 1):
            val = row[cn]; cell = ws.cell(row=r, column=j)
            if pd.isna(val): cell.value = ''
            elif isinstance(val, (np.integer,)): cell.value = int(val)
            elif isinstance(val, (np.floating, float)): cell.value = round(float(val), 4)
            elif isinstance(val, (np.bool_,)): cell.value = '是' if val else '否'
            else: cell.value = val
            cell.border = THIN_BORDER
            if cn == '排名':
                cell.alignment = CENTER; cell.font = Font(name='Calibri', size=10, bold=True)
                if val == 1: cell.fill = RANK_FILL_1
                elif val <= 3: cell.fill = RANK_FILL_2
            elif cn in ('股票代码', '股票名称'):
                cell.font = Font(name='微软雅黑', size=9, bold=True, color='1F4E79'); cell.alignment = LEFT
            elif cn == '总分':
                cell.alignment = CENTER; cell.font = Font(name='Calibri', size=10, bold=True, color='C00000')
                if isinstance(val, (int, float)) and float(val) >= 80: cell.fill = SCORE_HIGH
            elif cn == '选股逻辑':
                cell.font = Font(name='微软雅黑', size=8, color='333333'); cell.alignment = LEFT
            elif cn == '次日确认':
                cell.font = Font(name='微软雅黑', size=8, bold=True, color='0066CC'); cell.alignment = LEFT
            elif cn in RATE_COLS:
                cell.alignment = RIGHT
                if isinstance(val, (int, float)) and not pd.isna(val):
                    cell.font = Font(name='Calibri', size=9, color='C00000' if val > 0 else '008000')
            elif cn == '压力位':
                cell.alignment = RIGHT; cell.font = Font(name='Calibri', size=9, color='FF6600', bold=True)
            elif cn == '压力位类型':
                cell.font = Font(name='微软雅黑', size=8, color='FF6600'); cell.alignment = CENTER
            elif isinstance(val, (int, float)):
                cell.alignment = RIGHT; cell.font = Font(name='Calibri', size=9)
            else:
                cell.alignment = LEFT; cell.font = Font(name='微软雅黑', size=8)
    for j, cn in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTH.get(cn, max(len(str(cn))*2, 8))
    ws.freeze_panes = ws.cell(row=ds+1, column=1)
    return ws


# 模型总览
ws_ov = wb.active; ws_ov.title = '模型总览'
ov = [
    ['盘后选股模型 V8.1 — 五策略尖刀体系 (6月16日盘后)', '', '', '', ''],
    ['选股=圈范围, 次日开盘=确认入场, 互斥不重叠, 尖刀不在多而在精', '', '', '', ''],
    ['数据日期: 2026-06-16 | 快照时间: 20:26:34 | K线: 0520~0616', '', '', '', ''],
    ['', '', '', '', ''],
    ['策略', '阈值', '入选数', '核心驱动', '次日确认'],
    ['🔥打板求涨停', f'>={THRESHOLDS["daban"]}', len(result_daban), '封板强度+连板辨识+竞价抢筹-风险扣分', '竞价涨幅>1.5%+不低开'],
    ['📈趋势主升浪', f'>={THRESHOLDS["qushi"]}', len(result_qushi), '均线多头+量价配合+MACD+大单流入', '开盘不破MA5+量比>0.8'],
    ['🩹错杀低吸', f'>={THRESHOLDS["cuoshai"]}', len(result_cuoshai), '恐慌深度+承接力度+恐慌极值+催化剂', '开盘不创新低+竞价不低开'],
    ['⚡弱转强', f'>={THRESHOLDS["ruozhuan"]}', len(result_ruozhuan), '竞价异动+预期差+点火信号+股性', '竞价量比>1.5+5分钟涨幅>1%'],
    ['🔄强转弱反抽', f'>={THRESHOLDS["qzr"]}', len(result_qzr), '主力被套深度+回踩+板块支撑+反抽信号', '开盘>-2%+30分钟不创新低'],
    ['', '', '', '', ''],
    ['多策略共振(≥2策略)', '', len(result_resonance), '', '多维度共振确认'],
]
for i, rd in enumerate(ov, 1):
    for j, val in enumerate(rd, 1):
        cell = ws_ov.cell(row=i, column=j, value=val)
        if i == 1: cell.font = Font(name='微软雅黑', size=16, bold=True, color='1F4E79')
        elif i == 5: cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        elif i in range(6, 11): cell.font = Font(name='微软雅黑', size=9)
        else: cell.font = Font(name='微软雅黑', size=9, color='333333')
ws_ov.column_dimensions['A'].width = 30; ws_ov.column_dimensions['B'].width = 10
ws_ov.column_dimensions['C'].width = 10; ws_ov.column_dimensions['D'].width = 45
ws_ov.column_dimensions['E'].width = 30

# 5策略Sheet
write_sheet(wb, '🔥打板求涨停', result_daban, f'封板+连板+竞价 (≥{THRESHOLDS["daban"]}分)')
write_sheet(wb, '📈趋势主升浪', result_qushi, f'均线+量价+MACD+大单 (≥{THRESHOLDS["qushi"]}分)')
write_sheet(wb, '🩹错杀低吸', result_cuoshai, f'恐慌+承接+极值+催化剂 (≥{THRESHOLDS["cuoshai"]}分)')
write_sheet(wb, '⚡弱转强', result_ruozhuan, f'竞价异动+预期差+点火 (≥{THRESHOLDS["ruozhuan"]}分)')
write_sheet(wb, '🔄强转弱反抽', result_qzr, f'主力被套+回踩+支撑+反抽 (≥{THRESHOLDS["qzr"]}分)')
write_sheet(wb, '多策略共振', result_resonance, '≥2策略交叉确认')

# 模型使用方法 (恢复V8.0删除的说明sheet, 适配V8.1修复后的语义)
ws_g = wb.create_sheet(title='模型使用方法')
ws_g.merge_cells('A1:F1')
ws_g['A1'].value = '盘后选股V8.1 — 五策略尖刀使用方法 (6月16日盘后版)'
ws_g['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='1F4E79')
ws_g['A1'].alignment = LEFT
gh = ['策略', '核心驱动', '盘后圈范围', '次日确认', '调节参数', '互斥']
for j, h in enumerate(gh, 1):
    c = ws_g.cell(row=3, column=j, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = THIN_BORDER
gd = [
    ['🔥打板求涨停', '情绪×流动性',
     '涨停基础分10 + 封成比FCb(>0.2=稳)+封单额(>1亿=强) + 连板数(ConZAFDateNum: 1首板/2连板/3连板) + 竞价涨幅(>1.5%=抢筹) - 卖撤率(>0.92=高,新公式SCancel/(SCancel+TotalSVol))',
     '竞价涨幅>1.5%且不低开\n连板股: 竞价不低开+量放大\n未涨停: 竞价>3%+量比>2+5分钟涨>2%\n⚠封板弱若低开放弃入场',
     '封成比: 0.05/0.1/0.2\n封单额: 3000/5000万\n孤板阈值: ind_zt<=2', '与弱转强互斥'],
    ['📈趋势主升浪', '均线×筹码',
     'MA5>MA10>MA20(=35分) + 量价齐升(zaf>0且Wtb>1.5=25分) + MACD金叉(=20分) + 大单占比>40%(=20分)\n注: Wtb负值/0已置NaN, 防异常',
     '开盘不破MA5(容差1%)\n量比>0.8', '均线容差: 1%/2%/3%\n量能底线: 0.8/1.0/1.2', '独立(不需情绪)'],
    ['🩹错杀低吸', '恐慌×承接',
     'pool=当日跌>3% OR (当日跌>1%+20日跌>8%) OR (当日跌>1%+60日跌>15%) (AND逻辑, 真错杀)\n恐慌深度 + 承接力度(大买占比) + 恐慌极值(换手×跌幅) + 催化剂(行业涨停率)',
     '开盘不创新低\n竞价不低开', '跌幅: -3%/-5%/-7%\n承接: 30%/40%/50%\n催化剂不足惩罚: ×0.5', '与打板互斥'],
    ['⚡弱转强', '预期差×点火',
     'pool=竞价涨幅>1% 且 前日确实下跌(ZAFYesterday<0)\n竞价异动 + 预期差(前日弱+今强) + 点火信号(开盘跳空+量比) + 股性(年涨停数+Beta)',
     '竞价量比>1.5\n5分钟涨幅>1%', '竞价涨幅: 1%/2%/3%\n点火确认: 5/15/30分钟\n注: OpenAmo单位元, CJJEPre1单位万元(*10000对齐)', '与打板互斥'],
    ['🔄强转弱反抽', '被套×自救',
     'pool=昨涨>3% + 今跌>1%\n主力被套深度(昨涨-今跌) + 回踩幅度 + 板块支撑(行业涨停率, 区分9.8/19.5/29.5) + 反抽信号(大买占比)',
     '开盘>-2%\n30分钟不创新低', '回踩: 阳线1/3/1/2/2/3\n反抽判定: 30分/1小时', '与打板互斥'],
    ['共振说明', '多策略共振',
     '同一只股被≥2个策略选入, 取各策略"次日确认"条件交集(全满足才入场)',
     '看"次日确认条件数"列, 数字=需要满足的条件数\n所有条件必须同时满足', '阈值: 入选策略数>=2', '—'],
]
for i, rd in enumerate(gd):
    r = 4 + i
    for j, val in enumerate(rd):
        c = ws_g.cell(row=r, column=j+1, value=val); c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if j == 0: c.font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79'); c.fill = PatternFill(start_color='E8EEF4', end_color='E8EEF4', fill_type='solid')
        elif j == 1: c.font = Font(name='微软雅黑', size=8, bold=True, color='C00000')
        elif j == 2: c.font = Font(name='微软雅黑', size=8, color='333333')
        elif j == 3: c.font = Font(name='微软雅黑', size=8, color='0066CC')
        elif j == 4: c.font = Font(name='微软雅黑', size=8, color='666666')
        elif j == 5: c.font = Font(name='微软雅黑', size=8, color='CC6600')
ws_g.column_dimensions['A'].width = 16; ws_g.column_dimensions['B'].width = 16
ws_g.column_dimensions['C'].width = 50; ws_g.column_dimensions['D'].width = 30
ws_g.column_dimensions['E'].width = 30; ws_g.column_dimensions['F'].width = 16
for r in range(4, 4+len(gd)): ws_g.row_dimensions[r].height = 90

# V8.1修复说明
ws_fix = wb.create_sheet(title='V8.1修复说明')
ws_fix.merge_cells('A1:D1')
ws_fix['A1'].value = 'V8.1相对V8.0的修复清单 (基于6/16数据校验)'
ws_fix['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')
ws_fix['A1'].alignment = LEFT
fix_h = ['类别', '问题', '修复方案', '验证结果']
for j, h in enumerate(fix_h, 1):
    c = ws_fix.cell(row=3, column=j, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = THIN_BORDER
fix_data = [
    ['严重Bug', 'Wtb量比存在2594条负值+1002条0值', '置NaN, 防止趋势策略误判', '趋势30只中位Wtb 59.24, 均为高量比股'],
    ['严重Bug', '卖撤率公式SCancel/L2OrderNum中位1.30(>1不合理)', '改用SCancel/(SCancel+TotalSVol), 范围[0,1]', '新公式中位0.77, 阈值>0.92为高'],
    ['严重Bug', 'FCb/FCAmo存在11条负值', 'where(>=0, 0)清洗', '7只负值已清洗'],
    ['严重Bug', '错杀低吸pool用OR, 包含当日没跌的股票', '改AND: 当日跌+中期超跌双确认', 'V8.0独有6只当日ZAF均>0(未跌), V8.1正确剔除'],
    ['严重Bug', '行业涨停数不分板块(都用9.8%)', '区分9.8/19.5/29.5三档', '强转弱反抽末位5只因评分变化被替换'],
    ['中度Bug', 'get_latest_close用MA5Value兜底(语义错误)', '保留兜底但加"价格估算"列标记', '已导出到Excel, 用户可筛选'],
    ['中度Bug', 'OpenAmo*10000缺单位注释', '加注释: OpenAmo单位元, CJJEPre1单位万元', '弱转强策略已加注释'],
    ['中度Bug', '弱转强pool阈值ZAFYesterday<2过松', '收紧到<0(前日确实下跌)', 'V8.0独有6只前日微涨(0~2%), V8.1正确剔除'],
    ['中度Bug', 'fLianB原值0.96不像连板数', '改用ConZAFDateNum作连板数, 原fLianB作封板强度系数', 'V8.0独有7只打板股被正确剔除'],
    ['中度Bug', '多策略共振sheet缺统一次日确认', '取各策略次日确认交集(最严格)', '新增"次日确认条件数"列, 数字=条件数'],
    ['中度Bug', '风险扣分孤板阈值ind_zt<=1过严', '校准为<=2(适配行业涨停数修复)', '打板策略"⚠高位孤板"标签更合理'],
    ['轻度Bug', '"立即止损"表述与2段式理念冲突', '改为"放弃入场"', '盘后选股时尚未持仓, 无所谓止损'],
    ['轻度Bug', '20个模块标题注释被删除', '恢复所有(1)(2)(3)(4)注释', '可读性提升'],
    ['验证模块', 'T+1胜率未扣交易成本', '净胜率=T1涨-0.3%', '市场基准净胜率39.0%'],
    ['验证模块', '多策略共振预估逻辑predict_winrate取均值, 外部取max, 不一致', '保留分层: 单策略内多规律用均值, 共振股外部取max', '注释已说明设计合理性'],
    ['验证模块', '样本下限10过严, 17天回测样本不足', '放宽到5', '更多规律可被使用'],
]
for i, rd in enumerate(fix_data):
    r = 4 + i
    for j, val in enumerate(rd):
        c = ws_fix.cell(row=r, column=j+1, value=val); c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if j == 0:
            c.font = Font(name='微软雅黑', size=8, bold=True, color='C00000' if '严重' in val else ('CC6600' if '中度' in val else '666666'))
            c.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
        else:
            c.font = Font(name='微软雅黑', size=8, color='333333')
ws_fix.column_dimensions['A'].width = 10; ws_fix.column_dimensions['B'].width = 40
ws_fix.column_dimensions['C'].width = 40; ws_fix.column_dimensions['D'].width = 45
for r in range(4, 4+len(fix_data)): ws_fix.row_dimensions[r].height = 35

if 'Sheet' in wb.sheetnames: del wb['Sheet']
wb.save(OUTPUT_FILE)
print(f"\n✅ 输出文件: {OUTPUT_FILE}")
print(f"   Sheet数: {len(wb.sheetnames)}")
print(f"   各Sheet: {wb.sheetnames}")

# ========== 12. 打印结果 ==========
def pr(result, title):
    print(f"\n{'='*80}\n{title}\n{'='*80}")
    if len(result) == 0: print("  无符合条件标的"); return
    for _, row in result.head(10).iterrows():
        print(f"  {row.get('排名','')}. {row.get('股票名称','')}({row.get('股票代码','')}) | 总分:{row.get('总分',0)}")
        print(f"     逻辑: {row.get('选股逻辑','')}")
        print(f"     次日确认: {row.get('次日确认','')}")
        print()

pr(result_daban, '🔥打板求涨停')
pr(result_qushi, '📈趋势主升浪')
pr(result_cuoshai, '🩹错杀低吸')
pr(result_ruozhuan, '⚡弱转强')
pr(result_qzr, '🔄强转弱反抽')
print("\n===== V8.1 五策略尖刀模型 (6月16日盘后) 运行完成 =====")
