"""``/api/sectors`` 路由 - 板块管理。

- ``GET  /api/sectors``                       - 列出所有板块（从策略 YAML sector 段 + tqcenter 查询）
- ``GET  /api/sectors/export-all``            - 导出全部板块成份股（CSV / Excel）
- ``GET  /api/sectors/{code}/stocks``         - 获取板块成份股
- ``POST /api/sectors/{code}/refresh``        - 刷新板块（重新执行选股并回写）
- ``POST /api/sectors``                       - 占位（创建/更新板块，未实现具体逻辑）
"""

from __future__ import annotations

import io
import logging
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response

from engine.api.deps import get_config, get_runner, get_sector_manager, get_storage
from engine.api.schemas import (
    OkResponse,
    SectorInfoResponse,
    SectorRefreshResponse,
    SectorStockResponse,
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["sectors"])


@router.get(
    "",
    response_model=list[SectorInfoResponse],
    summary="列出所有板块",
)
async def list_sectors(
    cfg: Any = Depends(get_config),
    storage: Any = Depends(get_storage),
) -> list[SectorInfoResponse]:
    """合并两个来源：

    1. 策略 YAML 中 ``sector`` 段（``ZD_*`` 自定义板块，与策略一一对应）
    2. DuckDB ``sector_snapshots`` 表中最新的板块快照（含 stock_count）
    """
    snapshots = _query_latest_snapshots(storage)

    out: list[SectorInfoResponse] = []
    strategies = cfg.strategies() or {}
    seen: set[str] = set()
    for sid, sc in strategies.items():
        sector_obj = getattr(sc, "sector", None)
        if sector_obj is None:
            continue
        code = getattr(sector_obj, "code", "")
        if not code or code in seen:
            continue
        seen.add(code)
        snap = snapshots.get(code, {})
        out.append(
            SectorInfoResponse(
                code=code,
                name=getattr(sector_obj, "name", ""),
                strategy_id=sid,
                strategy_name=getattr(sc, "strategy_name", ""),
                stock_count=int(snap.get("stock_count", 0) or 0),
                auto_update=bool(getattr(sector_obj, "auto_update", True)),
                update_mode=str(getattr(sector_obj, "update_mode", "replace")),
                last_update=snap.get("snapshot_at"),
            )
        )
    out.sort(key=lambda s: s.code)
    return out


@router.post(
    "",
    response_model=OkResponse,
    summary="占位：创建/更新板块",
)
async def create_or_update_sector(
    sector_manager: Any = Depends(get_sector_manager),
) -> OkResponse:
    """前端 ``POST /api/sectors`` 的占位实现。

    实际创建需指定 code/name/stocks，本端点保留接口形态返回 ``{ok:true}``。
    """
    return OkResponse(ok=True, message="板块批量创建请走 /api/sectors/{code}/refresh")


@router.get(
    "/export-all",
    summary="导出全部板块成份股（CSV / Excel）",
    responses={
        200: {
            "content": {
                "text/csv": {},
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {},
            },
        },
    },
)
async def export_all_sectors(
    format: str = Query("csv", pattern="^(csv|excel)$"),
    cfg: Any = Depends(get_config),
    storage: Any = Depends(get_storage),
) -> Response:
    """导出所有板块的成份股到一个文件。

    - ``format=csv``: 多个板块用空行分隔, 每段含 ``# 板块: <code> <name> (<n> 只)`` 标题行
      列: ``stock_code, stock_name, score, added_at``
    - ``format=excel``: 多 Sheet 工作簿, 每个 Sheet 一个板块 (Sheet 名取板块 name 前 31 字符)
    """
    sectors: list[tuple[str, str]] = []  # (code, name)
    strategies = cfg.strategies() or {}
    seen: set[str] = set()
    for sid, sc in strategies.items():
        sector_obj = getattr(sc, "sector", None)
        if sector_obj is None:
            continue
        code = getattr(sector_obj, "code", "")
        if not code or code in seen:
            continue
        seen.add(code)
        name = getattr(sector_obj, "name", "") or code
        sectors.append((code, name))
    sectors.sort(key=lambda x: x[0])

    # 收集每个板块的成份股
    sector_stocks: list[tuple[str, str, list[SectorStockResponse]]] = []
    for code, name in sectors:
        stocks = _query_snapshot_stocks(storage, code)
        if not stocks:
            # 兜底: 跳过空板块 (不写入空段, 避免污染导出)
            continue
        sector_stocks.append((code, name, stocks))

    if not sector_stocks:
        raise HTTPException(status_code=404, detail="暂无板块数据可导出")

    # CSV
    if format == "csv":
        out = io.StringIO()
        out.write("\ufeff")  # BOM (Excel 兼容 UTF-8)
        for idx, (code, name, stocks) in enumerate(sector_stocks):
            if idx > 0:
                out.write("\n")
            out.write(f"# 板块: {code} {name} ({len(stocks)} 只)\n")
            out.write("stock_code,stock_name,score,added_at\n")
            for s in stocks:
                stock_name = (s.stock_name or "").replace(",", "_")
                added_at = (s.added_at or "").replace(",", " ")
                out.write(
                    f"{s.stock_code},{stock_name},{s.score:.3f},{added_at}\n"
                )
        content = out.getvalue().encode("utf-8")
        return Response(
            content=content,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": 'attachment; filename="sectors_all.csv"',
            },
        )

    # Excel (openpyxl)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="openpyxl 未安装，无法导出 Excel",
        ) from exc

    wb = Workbook()
    # 删除默认 Sheet
    default_ws = wb.active
    wb.remove(default_ws)

    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    headers = ["stock_code", "stock_name", "score", "added_at"]

    # 处理 Sheet 名重复 (openpyxl 不允许重名)
    used_names: set[str] = set()

    def _unique_sheet_name(raw: str) -> str:
        # Excel Sheet 名限制: 31 字符, 不能含 \ / ? * [ ]
        cleaned = "".join(c for c in raw if c not in "\\/?*[]:") or "Sheet"
        base = cleaned[:28] if len(cleaned) > 28 else cleaned
        candidate = base
        idx = 1
        while candidate in used_names:
            suffix = f"_{idx}"
            candidate = base[: 31 - len(suffix)] + suffix
            idx += 1
        used_names.add(candidate)
        return candidate

    for code, name, stocks in sector_stocks:
        ws = wb.create_sheet(title=_unique_sheet_name(name or code))
        # 表头
        for j, col in enumerate(headers, 1):
            cell = ws.cell(row=1, column=j, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        # 数据
        for i, s in enumerate(stocks, start=2):
            ws.cell(row=i, column=1, value=s.stock_code)
            ws.cell(row=i, column=2, value=s.stock_name or "")
            score_cell = ws.cell(row=i, column=3, value=round(float(s.score), 3))
            score_cell.alignment = center
            ws.cell(row=i, column=4, value=s.added_at or "")
        # 列宽
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 22
        ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="sectors_all.xlsx"',
        },
    )


@router.get(
    "/{code}/stocks",
    response_model=list[SectorStockResponse],
    summary="获取板块成份股",
)
async def get_sector_stocks(
    code: str,
    sector_manager: Any = Depends(get_sector_manager),
    storage: Any = Depends(get_storage),
) -> list[SectorStockResponse]:
    """优先查 ``sector_snapshots`` 表的最近快照，找不到再调适配器 ``get_user_sector``。"""
    snap_rows = _query_snapshot_stocks(storage, code)
    if snap_rows:
        return snap_rows

    # 兜底：调适配器
    try:
        stocks = sector_manager.get_stocks(code)
    except Exception as exc:  # noqa: BLE001
        logger.warning("get_stocks(%s) 失败: %s", code, exc)
        return []

    out: list[SectorStockResponse] = []
    for i, c in enumerate(stocks, start=1):
        out.append(
            SectorStockResponse(
                stock_code=c,
                stock_name="",
                added_at="",
                score=round(max(0.0, 1.0 - i * 0.02), 3),
            )
        )
    return out


@router.post(
    "/{code}/refresh",
    response_model=SectorRefreshResponse,
    summary="刷新板块成份股",
)
async def refresh_sector(
    code: str,
    cfg: Any = Depends(get_config),
    runner: Any = Depends(get_runner),
    sector_manager: Any = Depends(get_sector_manager),
    storage: Any = Depends(get_storage),
) -> SectorRefreshResponse:
    """重新执行该板块对应策略的选股，并把结果回写到通达信自定义板块。

    流程：
    1. 根据 ``code`` 反查策略 ID（从 ``cfg.strategies()`` 中匹配 ``sector.code``）
    2. 调 ``runner.run_strategy(sid)`` 触发选股（DuckDB / CSV / 板块导出由流水线自动完成）
    3. 把 ``ctx.final`` 中的 stock_code 列表调 ``sector_manager.update_stocks`` 原子回写
    """
    sid = _find_strategy_by_sector(cfg, code)
    if sid is None:
        raise HTTPException(
            status_code=404,
            detail=f"板块 {code} 未在 strategies/*.yaml 的 sector.code 中找到映射",
        )

    sc = cfg.strategy(sid)
    if sc is None:
        raise HTTPException(status_code=404, detail=f"策略 {sid} 不存在")
    if not getattr(sc, "enabled", True):
        raise HTTPException(status_code=400, detail=f"策略 {sid} 已禁用，无法刷新板块 {code}")

    try:
        ctx = runner.run_strategy(sid)
    except Exception as exc:  # noqa: BLE001
        logger.exception("刷新板块 %s 失败：策略 %s 执行异常", code, sid)
        return SectorRefreshResponse(
            ok=False, code=code, count=0, message=f"策略执行失败: {exc}"
        )

    # 提取选股结果中的代码列表
    stocks: list[str] = []
    if ctx.final is not None and not ctx.final.empty:
        for col in ("stock_code", "code", "Code"):
            if col in ctx.final.columns:
                stocks = [str(x) for x in ctx.final[col].tolist() if x]
                break

    if not stocks:
        return SectorRefreshResponse(
            ok=True, code=code, count=0, message=f"策略 {sid} 选出 0 只，板块未更新"
        )

    # 原子回写：clear + send_user_block（Mock 模式下 noop）
    try:
        ok = sector_manager.update_stocks(code, stocks)
    except Exception as exc:  # noqa: BLE001
        logger.exception("update_stocks(%s) 失败", code)
        return SectorRefreshResponse(
            ok=False, code=code, count=len(stocks), message=f"回写失败: {exc}"
        )

    return SectorRefreshResponse(
        ok=ok,
        code=code,
        count=len(stocks),
        message=f"策略 {sid} 选出 {len(stocks)} 只，已回写板块 {code}",
    )


# ============================================================================
# 内部
# ============================================================================


def _find_strategy_by_sector(cfg: Any, code: str) -> str | None:
    """根据板块代码反查 strategy_id。"""
    strategies = cfg.strategies() or {}
    for sid, sc in strategies.items():
        sector_obj = getattr(sc, "sector", None)
        if sector_obj and getattr(sector_obj, "code", "") == code:
            return sid
    return None


def _query_latest_snapshots(storage: Any) -> dict[str, dict[str, Any]]:
    """取每个 sector_code 最近一次 snapshot。"""
    if not _table_exists(storage, "sector_snapshots"):
        return {}
    sql = (
        "SELECT s.sector_code, s.sector_name, s.strategy_id, s.stock_count, "
        "       s.snapshot_at "
        "FROM sector_snapshots s "
        "INNER JOIN ("
        "  SELECT sector_code, MAX(snapshot_at) AS max_ts "
        "  FROM sector_snapshots GROUP BY sector_code"
        ") m ON s.sector_code = m.sector_code AND s.snapshot_at = m.max_ts"
    )
    try:
        df = storage.query(sql)
    except Exception as exc:  # noqa: BLE001
        logger.warning("查询 sector_snapshots 失败: %s", exc)
        return {}
    out: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        code = str(row.get("sector_code", ""))
        if not code:
            continue
        out[code] = {
            "stock_count": _to_int(row.get("stock_count")),
            "snapshot_at": _to_str(row.get("snapshot_at")),
        }
    return out


def _query_snapshot_stocks(storage: Any, code: str) -> list[SectorStockResponse]:
    """从最近一次 ``sector_snapshots`` 取 ``stock_list`` (JSON 数组)。

    增强：联表 ``selection_results`` 取股票中文名 + 真实得分（修复 stock_name 为空的 Bug）
    """
    import json

    if not _table_exists(storage, "sector_snapshots"):
        return []
    try:
        row = storage.fetchone(
            "SELECT stock_list, snapshot_at, strategy_id FROM sector_snapshots "
            "WHERE sector_code = ? ORDER BY snapshot_at DESC LIMIT 1",
            (code,),
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("查询 sector_snapshots(%s) 失败: %s", code, exc)
        return []
    if not row:
        return []
    raw_list = row[0]
    snap_at = _to_str(row[1]) if len(row) > 1 else None
    snap_strategy_id = str(row[2]) if len(row) > 2 and row[2] else ""
    try:
        codes = json.loads(str(raw_list)) if raw_list else []
    except (TypeError, ValueError, json.JSONDecodeError):
        return []

    # 联表查 selection_results 取 stock_name + total_score
    # 注: QuestDB 9.x 不支持 `run_id = (SELECT ... ORDER BY ... LIMIT 1)` 标量子查询比较
    # (报 "left operand must be a TIMESTAMP, found: SYMBOL")，改为两步查询避免子查询。
    code_to_info: dict[str, dict[str, Any]] = {}
    if codes and _table_exists(storage, "selection_results") and snap_strategy_id:
        try:
            # 第 1 步: 单独取该策略最近一次 run_id (ORDER BY created_at 单查可正常工作)
            run_df = storage.query(
                "SELECT run_id FROM selection_results "
                "WHERE strategy_id = ? ORDER BY created_at DESC LIMIT 1",
                [snap_strategy_id],
            )
            if run_df is not None and not run_df.empty:
                latest_run_id = str(run_df.iloc[0]["run_id"])
                # 第 2 步: 用 run_id + stock_code IN 取详情
                placeholders = ", ".join(["?"] * len(codes))
                sql = (
                    "SELECT stock_code, stock_name, total_score, rank "
                    "FROM selection_results "
                    f"WHERE strategy_id = ? AND stock_code IN ({placeholders}) AND run_id = ? "
                    "ORDER BY rank ASC"
                )
                params = [snap_strategy_id, *codes, latest_run_id]
                df = storage.query(sql, params)
                for _, r in df.iterrows():
                    code_to_info[str(r["stock_code"])] = {
                        "name": str(r.get("stock_name", "") or ""),
                        "score": float(r.get("total_score") or 0.0),
                        "rank": int(r.get("rank") or 0),
                    }
        except Exception as exc:  # noqa: BLE001
            logger.warning("联表查 selection_results 失败: %s", exc)

    out: list[SectorStockResponse] = []
    for c in codes:
        info = code_to_info.get(str(c), {})
        out.append(
            SectorStockResponse(
                stock_code=str(c),
                stock_name=str(info.get("name", "")),
                added_at=snap_at or "",
                # 联表查到真实 total_score 用之；查不到为 0.0（诚实），不再造假分 100-i*2
                score=float(info.get("score", 0.0)),
            )
        )
    return out


def _table_exists(storage: Any, name: str) -> bool:
    if storage is None:
        return False
    try:
        return storage.table_exists(name)
    except Exception:  # noqa: BLE001
        return False


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    try:
        if hasattr(v, "isoformat"):
            return v.isoformat()
        return str(v)
    except Exception:  # noqa: BLE001
        return None


def _to_int(v: Any) -> int:
    try:
        if v is None or (isinstance(v, float) and v != v):
            return 0
        return int(v)
    except (TypeError, ValueError):
        return 0
