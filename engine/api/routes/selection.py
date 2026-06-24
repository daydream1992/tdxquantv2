"""``/api/selections`` 路由 - 选股结果查询与导出。

- ``GET  /api/selections``                       - 列表（支持 strategy_id/date/min_score 等筛选）
- ``GET  /api/selections/{run_id}``              - 单次结果详情
- ``GET  /api/selections/{run_id}/export``       - 导出 CSV / Excel
"""

from __future__ import annotations

import io
import json
import logging
from typing import Any

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response

from engine.api.deps import get_config, get_storage
from engine.api.schemas import (
    SelectionDetailResponse,
    SelectionFactorScore,
    SelectionRowResponse,
)

logger = logging.getLogger(__name__)


def _safe_float(v: Any, default: float = 0.0) -> float:
    """float 转换，容忍 None / np.nan。

    QuestDBStore 出口已把 np.nan 清成 None，此处为 JSON 序列化前的防御性守卫：
    ``bool(np.nan) is True`` 会使 ``x or default`` 失效，导致 NaN 泄漏进响应触发 500。
    """
    if v is None or pd.isna(v):
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _safe_int(v: Any, default: int = 0) -> int:
    """int 转换，容忍 None / np.nan（``int(np.nan)`` 会抛 ValueError）。"""
    if v is None or pd.isna(v):
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


router = APIRouter(tags=["selection"])


# ============================================================================
# 列表
# ============================================================================


@router.get(
    "",
    response_model=list[SelectionRowResponse],
    summary="列出选股结果",
)
async def list_selections(
    storage: Any = Depends(get_storage),
    cfg: Any = Depends(get_config),
    strategy_id: str | None = Query(None, description="按策略 ID 筛选"),
    run_id: str | None = Query(None, description="按 run_id 筛选"),
    start_date: str | None = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: str | None = Query(None, description="结束日期 YYYY-MM-DD"),
    min_score: float | None = Query(None, ge=0, le=100, description="最低得分门槛（0-100）"),
    limit: int = Query(200, ge=1, le=2000),
) -> list[SelectionRowResponse]:
    """从 DuckDB ``selection_results`` 表查询。"""
    if not _table_exists(storage, "selection_results"):
        return []

    where_parts: list[str] = []
    params: list[Any] = []
    if strategy_id:
        where_parts.append("strategy_id = ?")
        params.append(strategy_id)
    if run_id:
        where_parts.append("run_id = ?")
        params.append(run_id)
    if min_score is not None:
        where_parts.append("total_score >= ?")
        params.append(float(min_score))
    if start_date:
        where_parts.append("run_date >= ?")
        params.append(start_date)
    if end_date:
        where_parts.append("run_date <= ?")
        params.append(end_date)
    where_clause = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""

    sql = (
        "SELECT run_id, strategy_id, run_date, stock_code, stock_name, "
        "       total_score, factor_scores, rank, created_at "
        f"FROM selection_results{where_clause} "
        "ORDER BY run_date DESC, rank ASC "
        "LIMIT ?"
    )
    params.append(int(limit))

    try:
        df = storage.query(sql, params)
    except Exception as exc:  # noqa: BLE001
        logger.warning("查询 selection_results 失败: %s", exc)
        return []

    return _rows_from_df(df, cfg)


# ============================================================================
# 详情
# ============================================================================


@router.get(
    "/{run_id}",
    response_model=SelectionDetailResponse,
    summary="获取单次选股结果详情",
)
async def get_selection_detail(
    run_id: str,
    storage: Any = Depends(get_storage),
    cfg: Any = Depends(get_config),
) -> SelectionDetailResponse:
    """返回某次 run 的全部股票 + 该 run 的元信息。"""
    if not _table_exists(storage, "selection_results"):
        raise HTTPException(status_code=404, detail=f"run_id={run_id} 不存在（无 selection_results 表）")
    sql = (
        "SELECT run_id, strategy_id, run_date, stock_code, stock_name, "
        "       total_score, factor_scores, rank, created_at "
        "FROM selection_results WHERE run_id = ? "
        "ORDER BY rank ASC"
    )
    try:
        df = storage.query(sql, (run_id,))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"查询失败: {exc}") from exc

    if df.empty:
        raise HTTPException(status_code=404, detail=f"run_id={run_id} 无数据")

    first = df.iloc[0]
    rows = _rows_from_df(df, cfg)
    strategy_id_str = str(first.get("strategy_id", ""))
    strategy_name = _lookup_strategy_name(cfg, strategy_id_str)

    # 取 strategy_runs 元信息（如果表存在）
    started_at: str | None = None
    finished_at: str | None = None
    duration_sec: float | None = None
    status = "success"
    if _table_exists(storage, "strategy_runs"):
        try:
            run_df = storage.query(
                "SELECT started_at, finished_at, duration_ms, status "
                "FROM strategy_runs WHERE run_id = ?",
                (run_id,),
            )
            if not run_df.empty:
                r0 = run_df.iloc[0]
                started_at = _to_str(r0.get("started_at"))
                finished_at = _to_str(r0.get("finished_at"))
                if r0.get("duration_ms") is not None:
                    try:
                        duration_sec = round(float(r0["duration_ms"]) / 1000.0, 3)
                    except (TypeError, ValueError):
                        duration_sec = None
                status = str(r0.get("status", "success"))
        except Exception as exc:  # noqa: BLE001
            logger.warning("查询 strategy_runs 失败: %s", exc)

    return SelectionDetailResponse(
        run_id=run_id,
        strategy_id=strategy_id_str,
        strategy_name=strategy_name,
        started_at=started_at,
        finished_at=finished_at,
        duration_sec=duration_sec,
        n_stocks=len(rows),
        status=status,
        rows=rows,
    )


# ============================================================================
# 导出
# ============================================================================


@router.get(
    "/{run_id}/export",
    summary="导出选股结果（CSV 或 Excel）",
    responses={
        200: {"content": {"text/csv": {}, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}}},
    },
)
async def export_selection(
    run_id: str,
    storage: Any = Depends(get_storage),
    format: str = Query("csv", pattern="^(csv|excel)$"),
) -> Response:
    """导出某次 run 的全部结果。

    - ``format=csv``    -> ``text/csv; charset=utf-8`` (BOM 头，Excel 兼容)
    - ``format=excel``  -> ``.xlsx`` 二进制流
    """
    if not _table_exists(storage, "selection_results"):
        raise HTTPException(status_code=404, detail=f"run_id={run_id} 无数据")

    sql = (
        "SELECT run_id, strategy_id, run_date, stock_code, stock_name, "
        "       total_score, factor_scores, rank, created_at "
        "FROM selection_results WHERE run_id = ? "
        "ORDER BY rank ASC"
    )
    try:
        df = storage.query(sql, (run_id,))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"查询失败: {exc}") from exc

    if df.empty:
        raise HTTPException(status_code=404, detail=f"run_id={run_id} 无数据")

    export_df = _format_export_df(df)

    if format == "csv":
        buf = io.StringIO()
        export_df.to_csv(buf, index=False, encoding="utf-8-sig")
        content = "\ufeff" + buf.getvalue()  # BOM
        return Response(
            content=content.encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="selection_{run_id}.csv"',
            },
        )

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="openpyxl 未安装，无法导出 Excel",
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "选股结果"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    cols = list(export_df.columns)
    for j, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=str(col))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for i, (_, row) in enumerate(export_df.iterrows(), start=2):
        for j, col in enumerate(cols, 1):
            val = row[col]
            if pd.isna(val):
                val = ""
            ws.cell(row=i, column=j, value=val if not isinstance(val, pd.Timestamp) else val.isoformat())

    bio = io.BytesIO()
    wb.save(bio)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="selection_{run_id}.xlsx"',
        },
    )


# ============================================================================
# 内部辅助
# ============================================================================


def _rows_from_df(df: pd.DataFrame, cfg: Any = None) -> list[SelectionRowResponse]:
    """把 DataFrame 转 SelectionRowResponse 列表。

    增强：
    - ``strategy_name`` 从 ``cfg.strategies()`` 反查填充（之前为空）
    - ``factors[].weight`` 从策略 YAML 的 ``factors`` 段反查填充（之前为 0）
    """
    # 预构建 strategy_id → (name, factor_id → weight) 映射
    strategy_map: dict[str, dict[str, Any]] = {}
    if cfg is not None:
        try:
            strategies = cfg.strategies() or {}
            for sid, sc in strategies.items():
                factors_dict: dict[str, float] = {}
                for f in getattr(sc, "factors", []) or []:
                    fid = getattr(f, "factor_id", "")
                    w = float(getattr(f, "weight", 1.0))
                    if fid:
                        factors_dict[fid] = w
                strategy_map[sid] = {
                    "name": getattr(sc, "strategy_name", "") or sid,
                    "factors": factors_dict,
                }
        except Exception as exc:  # noqa: BLE001
            logger.warning("构建 strategy_map 失败: %s", exc)

    out: list[SelectionRowResponse] = []
    for _, row in df.iterrows():
        sid = str(row.get("strategy_id", ""))
        sm = strategy_map.get(sid, {})
        strategy_name = sm.get("name", "") or sid
        factor_weights = sm.get("factors", {})

        factors_raw = row.get("factor_scores", "{}")
        factors_list = _parse_factor_scores(factors_raw, factor_weights)
        out.append(
            SelectionRowResponse(
                run_id=str(row.get("run_id", "")),
                strategy_id=sid,
                strategy_name=strategy_name,
                stock_code=str(row.get("stock_code", "")),
                stock_name=str(row.get("stock_name", "")),
                score=_safe_float(row.get("total_score")),
                rank=_safe_int(row.get("rank")),
                factors=factors_list,
                run_at=_to_str(row.get("created_at")) or _to_str(row.get("run_date")) or "",
            )
        )
    return out


def _lookup_strategy_name(cfg: Any, strategy_id: str) -> str:
    """从 ConfigLoader 反查策略中文名。"""
    if not strategy_id or cfg is None:
        return ""
    try:
        sc = cfg.strategy(strategy_id)
        if sc is not None:
            return getattr(sc, "strategy_name", "") or strategy_id
    except Exception:  # noqa: BLE001
        pass
    return strategy_id


def _parse_factor_scores(raw: Any, factor_weights: dict[str, float] | None = None) -> list[SelectionFactorScore]:
    """``factor_scores`` 列存的是 JSON 字符串 ``{"factor_id": score}``。

    Args:
        raw: DuckDB 中存的 JSON 字符串 / dict
        factor_weights: 策略 YAML 中各因子的权重映射 ``{factor_id: weight}``，
            用于填充 ``SelectionFactorScore.weight`` 字段（DuckDB 中不存权重）
    """
    if not raw:
        return []
    if isinstance(raw, dict):
        d = raw
    else:
        try:
            d = json.loads(str(raw))
        except (TypeError, ValueError, json.JSONDecodeError):
            return []
    if not isinstance(d, dict):
        return []
    weights = factor_weights or {}
    out: list[SelectionFactorScore] = []
    for k, v in d.items():
        try:
            val = float(v) if v is not None else 0.0
        except (TypeError, ValueError):
            val = 0.0
        out.append(
            SelectionFactorScore(
                factor_id=str(k),
                value=val,
                weight=float(weights.get(str(k), 0.0)),
                score=val,
            )
        )
    return out


def _format_export_df(df: pd.DataFrame) -> pd.DataFrame:
    """整理导出列：选择/重命名/丢弃 JSON 字段。"""
    cols = [
        "run_id", "strategy_id", "run_date", "stock_code", "stock_name",
        "total_score", "rank", "created_at",
    ]
    keep = [c for c in cols if c in df.columns]
    out = df[keep].copy()
    # 重命名为中文表头（与前端 mock-data 字段对齐）
    rename_map = {
        "total_score": "score",
        "created_at": "run_at",
        "run_date": "run_date",
    }
    out = out.rename(columns={k: v for k, v in rename_map.items() if k in out.columns})
    return out


def _table_exists(storage: Any, name: str) -> bool:
    if storage is None:
        return False
    try:
        return storage.table_exists(name)
    except Exception:  # noqa: BLE001
        return False


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    try:
        if hasattr(v, "isoformat"):
            return v.isoformat()
        return str(v)
    except Exception:  # noqa: BLE001
        return None
