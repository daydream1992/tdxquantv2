"""Excel 多 Sheet 导出器（V8 兼容）。

V8 输出 Excel 包含 8 个 Sheet:
- 5 个策略 Sheet (打板求涨停/趋势主升浪/错杀低吸/弱转强/强转弱反抽)
- 多策略共振 Sheet
- 汇总 Sheet
- 日志 Sheet

本导出器设计:
- 单策略运行时: 只写当前策略 + 日志 Sheet
- 全量运行时: 读 ``config/export.yaml`` 的 ``excel.sheets`` 列表，依次写入
- 样式: 沿用 V8 配色（深蓝表头/排名高亮/红涨绿跌）

Sheet 配置示例 (``config/export.yaml`` ``excel.sheets``):
.. code-block:: yaml
    sheets:
      - sheet_name: 打板求涨停
        strategy_id: dbqzt
      - sheet_name: 多策略共振
        strategy_id: _resonance
      - sheet_name: 汇总
        strategy_id: _summary
      - sheet_name: 日志
        strategy_id: _log
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

from engine.exporters.base import DataExporter, ExporterError
from engine.pipeline.base import PipelineContext

logger = logging.getLogger(__name__)

# 兼容性: openpyxl 必须可用
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_READY = True
except ImportError:  # pragma: no cover
    _OPENPYXL_READY = False
    Workbook = None  # type: ignore


# V8 配色（与 V8 run.py 完全一致）
HEADER_FILL_COLOR = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"
RANK1_FILL_COLOR = "FFF2CC"
RANK3_FILL_COLOR = "E2EFDA"
SCORE_HIGH_FILL_COLOR = "FCE4EC"
BORDER_COLOR = "D9D9D9"
THIN_BORDER_STYLE = "thin"


class ExcelExporter(DataExporter):
    """Excel 导出器。"""

    exporter_id = "excel"
    exporter_name = "Excel 导出"

    def export(self, context: PipelineContext) -> str:
        if not _OPENPYXL_READY:
            raise ExporterError("openpyxl 未安装，无法导出 Excel")

        output_dir = Path(self.config.get("output_dir", "./data/excel"))
        output_dir.mkdir(parents=True, exist_ok=True)

        pattern = self.config.get("filename_pattern", "{date}_selection.xlsx")
        today = datetime.now().strftime("%Y%m%d")
        filename = pattern.format(date=today, strategy_id=context.strategy_id, run_id=context.run_id)
        path = output_dir / filename

        # 准备 sheets 配置
        sheets_cfg: list[dict[str, Any]] = self.config.get("sheets", []) or []
        # 过滤启用项
        enabled_sheets = [s for s in sheets_cfg if s.get("enabled", True)]

        # 始终至少有当前策略 sheet + 日志 sheet
        if not enabled_sheets:
            enabled_sheets = [
                {"sheet_name": context.config.get("strategy_name", context.strategy_id),
                 "strategy_id": context.strategy_id},
                {"sheet_name": "日志", "strategy_id": "_log"},
            ]

        wb = Workbook()
        wb.properties.creator = "TdxQuant Engine"

        first = True
        for sheet_cfg in enabled_sheets:
            sheet_name = sheet_cfg.get("sheet_name", "Sheet")
            strategy_id = sheet_cfg.get("strategy_id", "")
            df = self._select_sheet_data(strategy_id, context)
            desc = sheet_cfg.get("description", "")
            ws = self._write_sheet(wb, sheet_name, df, desc, context, first=first)
            first = False

        # 删除默认 Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(path)
        self.logger.info("Excel 导出: %s (%d 个 Sheet)", path, len(wb.sheetnames))
        return str(path)

    # ---- Sheet 数据选择 ----
    def _select_sheet_data(self, strategy_id: str, context: PipelineContext) -> pd.DataFrame:
        """根据 sheet 配置的 strategy_id 选择数据。

        特殊 strategy_id:
        - ``_log``     - 写入运行日志（步骤耗时/警告）
        - ``_summary`` - 汇总（各策略入选数）
        - ``_resonance``- 多策略共振（P1-5 实现，目前返回空）
        - 空字符串     - 与当前策略相同
        """
        if strategy_id == "_log":
            return self._build_log_df(context)
        if strategy_id == "_summary":
            return self._build_summary_df(context)
        if strategy_id == "_resonance":
            # TODO(P1-5): 多策略共振需跨策略聚合，本阶段返回空
            return pd.DataFrame(columns=["提示"])
        # 普通策略 sheet
        if strategy_id and strategy_id != context.strategy_id:
            # 其他策略 sheet: 当前 context 没有，返回空提示
            return pd.DataFrame(columns=[f"策略 {strategy_id} 未运行"])
        return context.final if context.final is not None else pd.DataFrame()

    def _build_log_df(self, context: PipelineContext) -> pd.DataFrame:
        rows = []
        for step_info in context.metadata.get("steps", []):
            rows.append({
                "步骤": step_info.get("step_id"),
                "状态": step_info.get("status"),
                "耗时(秒)": round(step_info.get("duration_sec", 0), 3),
                "错误": step_info.get("error", ""),
            })
        warnings = context.metadata.get("warnings", [])
        for w in warnings:
            rows.append({
                "步骤": w.get("step_id"),
                "状态": "warning",
                "耗时(秒)": "",
                "错误": w.get("message"),
            })
        if not rows:
            rows.append({"步骤": "无", "状态": "ok", "耗时(秒)": 0, "错误": ""})
        return pd.DataFrame(rows)

    def _build_summary_df(self, context: PipelineContext) -> pd.DataFrame:
        n_final = 0 if context.final is None else len(context.final)
        return pd.DataFrame([
            {"策略": context.config.get("strategy_name", context.strategy_id),
             "策略ID": context.strategy_id,
             "run_id": context.run_id,
             "入选数": n_final,
             "开始时间": context.started_at.isoformat(),
             "耗时(秒)": round(context.duration_sec or 0, 3)},
        ])

    # ---- Sheet 写入（V8 风格） ----
    def _write_sheet(
        self,
        wb: Any,
        sheet_name: str,
        df: pd.DataFrame,
        desc: str,
        context: PipelineContext,
        first: bool = False,
    ) -> Any:
        ws = wb.active if first else wb.create_sheet(title=sheet_name)
        if first:
            ws.title = sheet_name

        # 标题行
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"{sheet_name} — {desc}" if desc else sheet_name
        title_cell.font = Font(name="微软雅黑", size=14, bold=True, color=HEADER_FILL_COLOR)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # 副标题
        ws.merge_cells("A2:H2")
        subtitle = (
            f"策略: {context.strategy_id} | run_id: {context.run_id} | "
            f"开始: {context.started_at.isoformat()} | 共 {len(df)} 只入选"
        )
        sub_cell = ws["A2"]
        sub_cell.value = subtitle
        sub_cell.font = Font(name="微软雅黑", size=9, color="666666")
        sub_cell.alignment = Alignment(horizontal="left", vertical="center")

        if df.empty:
            ws.cell(row=3, column=1, value="无符合条件标的").font = Font(
                name="微软雅黑", size=11, color="999999"
            )
            return ws

        # 表头
        cols = list(df.columns)
        header_row = 3
        header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
        header_font = Font(name="微软雅黑", size=10, bold=True, color=HEADER_FONT_COLOR)
        thin_side = Side(style=THIN_BORDER_STYLE, color=BORDER_COLOR)
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for j, col_name in enumerate(cols, 1):
            cell = ws.cell(row=header_row, column=j, value=str(col_name))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

        # 数据行
        rank1_fill = PatternFill(start_color=RANK1_FILL_COLOR, end_color=RANK1_FILL_COLOR, fill_type="solid")
        rank3_fill = PatternFill(start_color=RANK3_FILL_COLOR, end_color=RANK3_FILL_COLOR, fill_type="solid")
        score_high_fill = PatternFill(start_color=SCORE_HIGH_FILL_COLOR, end_color=SCORE_HIGH_FILL_COLOR, fill_type="solid")

        for i, (_, row) in enumerate(df.iterrows()):
            r = header_row + 1 + i
            for j, col_name in enumerate(cols, 1):
                val = row[col_name]
                cell = ws.cell(row=r, column=j)
                if pd.isna(val):
                    cell.value = ""
                elif isinstance(val, bool):
                    cell.value = "是" if val else "否"
                elif isinstance(val, (int,)) and not isinstance(val, bool):
                    cell.value = int(val)
                elif isinstance(val, (float,)):
                    cell.value = round(float(val), 4)
                else:
                    cell.value = val
                cell.border = thin_border

                if col_name in ("rank", "排名"):
                    cell.alignment = center
                    cell.font = Font(name="Calibri", size=10, bold=True)
                    if val == 1:
                        cell.fill = rank1_fill
                    elif isinstance(val, (int, float)) and val <= 3:
                        cell.fill = rank3_fill
                elif col_name in ("total_score", "总分"):
                    cell.alignment = center
                    cell.font = Font(name="Calibri", size=10, bold=True, color="C00000")
                    if isinstance(val, (int, float)) and float(val) >= 80:
                        cell.fill = score_high_fill

        # 列宽
        for j, col_name in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(j)].width = max(
                len(str(col_name)) * 2, 8, 10
            )

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        return ws
