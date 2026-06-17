#!/usr/bin/env python3
"""
把T+1预估验证结果合并到主选股Excel中
- 主Excel: 盘后选股模型_V8_1_20260616_盘后.xlsx
- 验证Excel: V8_1_0616_T+1预估验证.xlsx
- 合并后主Excel新增一个Sheet: T+1预估验证
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAIN_FILE = 'output/盘后选股模型_V8_1_20260616_盘后.xlsx'
VERIFY_FILE = 'output/V8_1_0616_T+1预估验证.xlsx'

# 读取T+1预估验证数据
verify_df = pd.read_excel(VERIFY_FILE)
print(f"读取T+1预估验证: {len(verify_df)}条")

# 按预估胜率排序, 让高分股在前
verify_df = verify_df.sort_values(['评级','预估胜率%'], ascending=[True, False])

# 加载主Excel
wb = load_workbook(MAIN_FILE)
print(f"主Excel当前Sheet: {wb.sheetnames}")

# 如果已存在T+1预估验证Sheet, 删除重建
if 'T+1预估验证' in wb.sheetnames:
    del wb['T+1预估验证']

# 创建新Sheet
ws = wb.create_sheet(title='T+1预估验证', index=1)  # 插入到第2位(模型总览后)

# 样式
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
RANK_FILL_1 = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
RANK_FILL_2 = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
STAR3_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
STAR2_FILL = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                     top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

# 标题行
ws.merge_cells('A1:J1')
ws['A1'].value = 'T+1预估验证 — 6月16日盘后选股次日表现预估'
ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')
ws['A1'].alignment = LEFT

ws.merge_cells('A2:J2')
ws['A2'].value = '基于历史17天(0521~0612)45条硬规律外推 | 评级: ★★★(≥60%) / ★★(55-60%) / ★(50-55%) / △(<50%) | 共 133 只'
ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
ws['A2'].alignment = LEFT

# 表头
cols = list(verify_df.columns)
ds = 3
for j, cn in enumerate(cols, 1):
    cell = ws.cell(row=ds, column=j, value=cn)
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = THIN_BORDER

# 数据行
for i, (_, row) in enumerate(verify_df.iterrows()):
    r = ds + 1 + i
    for j, cn in enumerate(cols, 1):
        val = row[cn]
        cell = ws.cell(row=r, column=j)
        if pd.isna(val):
            cell.value = ''
        elif isinstance(val, (int, float)) and not isinstance(val, bool):
            cell.value = round(float(val), 3)
            cell.alignment = RIGHT
            cell.font = Font(name='Calibri', size=9)
            # 预估胜率列特殊着色
            if cn == '预估胜率%':
                if val >= 55:
                    cell.font = Font(name='Calibri', size=9, bold=True, color='C00000')
                elif val >= 50:
                    cell.font = Font(name='Calibri', size=9, bold=True, color='FF6600')
                else:
                    cell.font = Font(name='Calibri', size=9, color='666666')
            elif cn == '预估均涨%':
                if val > 0:
                    cell.font = Font(name='Calibri', size=9, color='C00000')
                else:
                    cell.font = Font(name='Calibri', size=9, color='008000')
            elif cn == '涨幅%':
                if val > 0:
                    cell.font = Font(name='Calibri', size=9, color='C00000')
                else:
                    cell.font = Font(name='Calibri', size=9, color='008000')
        else:
            cell.value = str(val) if val is not None else ''
            cell.alignment = LEFT
            cell.font = Font(name='微软雅黑', size=8)
        cell.border = THIN_BORDER
        
        # 评级列特殊样式
        if cn == '评级':
            cell.alignment = CENTER
            if val == '★★★':
                cell.font = Font(name='微软雅黑', size=11, bold=True, color='C00000')
                cell.fill = STAR3_FILL
            elif val == '★★':
                cell.font = Font(name='微软雅黑', size=11, bold=True, color='FF6600')
                cell.fill = STAR2_FILL
            elif val == '★':
                cell.font = Font(name='微软雅黑', size=10, color='CC9900')
        # 入选策略列特殊样式
        elif cn == '入选策略':
            cell.alignment = CENTER
            cell.font = Font(name='微软雅黑', size=9, bold=True, color='1F4E79')
        # 股票代码/名称
        elif cn in ('股票代码', '股票名称'):
            cell.font = Font(name='微软雅黑', size=9, bold=True, color='1F4E79')
            cell.alignment = LEFT

# 列宽
COL_WIDTH = {
    '股票代码': 12, '股票名称': 12, '入选策略': 14, '涨幅%': 8, '总分': 8,
    '预估胜率%': 12, '预估均涨%': 12, '评级': 8, '匹配规律': 50, '选股逻辑': 50,
}
for j, cn in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(j)].width = COL_WIDTH.get(cn, max(len(str(cn))*2, 10))

ws.freeze_panes = ws.cell(row=ds+1, column=1)

# 保存
wb.save(MAIN_FILE)
print(f"\n✅ 已合并到主Excel")
print(f"   文件: {MAIN_FILE}")
print(f"   Sheet: {wb.sheetnames}")
print(f"   T+1预估验证行数: {len(verify_df)}")
